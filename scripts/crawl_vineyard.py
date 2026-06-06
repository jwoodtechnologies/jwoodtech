"""
================================================================================
VINEYARD CITY GOVERNMENT RESEARCH TOOL — STANDALONE CRAWL SCRIPT
================================================================================

AUDIT REPORT
============

Infrastructure Overview
-----------------------
The production backend at backend/server.py (~3430 lines) is a FastAPI
application deployed on Render. It uses Motor (asyncio MongoDB driver) connected
to MongoDB Atlas. The Vercel function at api/index.py loads a lighter-weight
vineyard_router.py for auth/search (no crawling), while all crawling remains
exclusively in server.py.

MongoDB Atlas Configuration
----------------------------
- Connection: MONGO_URL env var (motor AsyncIOMotorClient)
- Database:   DB_NAME env var (defaults to "jwoodtech" in vercel entry point)
- No .env file found in backend/ at audit time — credentials live only in the
  host environment (Render env vars for production, local shell for scripts).

Collections Present in Code
-----------------------------
  db.documents      — the primary search index; each row is a text chunk
  db.sources        — one row per crawl source (URL, label, status, metrics)
  db.index_meta     — version pointer: kind=active/build/previous
  db.metrics        — atomic counters (search count, OpenAI token spend, etc.)
  db.contact_submissions — contact form leads
  db.chatbot_submissions — chatbot leads

Document Schema (db.documents)
--------------------------------
  id, source_id, source_root, source_label, source_site, url, pdf_url,
  title, section_ref, excerpt, content, is_pdf, is_xlsx, doc_type,
  meeting_date, depth, index_version, crawl_run_id (transient), created_at

  source_site buckets: civicclerk | municode | vineyardutah | rda | other
  doc_type values:     resolution | ordinance | minutes | agenda |
                       attachment | transparency | rda | pdf | xlsx | page

Index Versioning Scheme
------------------------
  Crawls write documents tagged index_version = "build-<run_id>".
  User-facing reads are scoped to the version stored in index_meta{kind=active}.
  Admin promotes a build version via /vineyard/admin/lock-index.
  One previous version is retained as prev-<ver> for rollback.
  This script replicates that exact scheme.

Content Targeted
-----------------
  1. CivicClerk (vineyardut.portal.civicclerk.com)
     - OData REST API: /v1/Events (paginated, both asc and desc direction)
     - Per-event: /v1/Meetings/{agendaId} for published files + agenda outline
     - Files fetched as plainText=true stream; PyMuPDF fallback for older PDFs
     - Covers: agendas, minutes, agenda packets, attachments, resolutions

  2. Municipal Code Online (vineyard.municipalcodeonline.com)
     - AngularJS SPA — requires Playwright (headless Chromium)
     - Intercepts XHR: /book/expand (tree), /book/content (section text)
     - Walks 9 MCO_TYPES: ordinances, districts, resolutions, plan,
       subdivords, orddoc, minutes, zoning, landscaping
     - Granular section split via phx-name HTML markers

  3. Vineyard Utah Official Site (www.vineyardutah.gov)
     - Generic BFS crawler (httpx + BeautifulSoup)
     - Handles HTML pages, PDF documents, XLSX spreadsheets
     - Recursive within-host, depth-limited (MAX_DEPTH=10)
     - Respects MAX_PAGES_PER_SOURCE (default 5000)

  4. RDA Past Meetings XLSX (revize.com) — legacy, hidden in UI
     - Seeded as a source but typically omitted from active chips

What Is Working
----------------
  - Full Atlas persistence with index versioning
  - CivicClerk OData crawler is comprehensive (all events, all files)
  - MCO Playwright crawler handles the SPA architecture correctly
  - Generic crawler handles HTML/PDF/XLSX with dedup
  - BM25 + OpenAI embedding hybrid search
  - Index lock/rollback admin controls

Potential Gaps / Attention Items
----------------------------------
  - No .env file in backend/ — credentials must be supplied via environment
  - Playwright requires `playwright install chromium` before MCO crawl works
  - The Atlas free tier (512 MB) comment in code notes embeddings are skipped
    to conserve storage; BM25 search still works but semantic search is limited
  - CRAWL_TIMEOUT_SECONDS defaults to 1800s (30 min); MCO crawl can be slow
    on a cold Playwright start
  - The RDA XLSX URL (revize.com) may return 404 or redirect; the generic
    crawler will skip it gracefully
  - No MongoDB text index is created by the application code — the text index
    on db.documents (title + content fields) must be created manually in Atlas
    or by running this script with --ensure-indexes (not yet implemented here).
    Without it, the $text search prefilter is silently skipped and the code
    falls back to a full collection scan (slower but functional).

================================================================================
"""
from __future__ import annotations

import argparse
import asyncio
import collections
import io
import json
import logging
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional
from urllib.parse import parse_qs, urljoin, urldefrag, urlparse

# ---------------------------------------------------------------------------
# Environment — load backend/.env if present, then fall back to os.environ
# ---------------------------------------------------------------------------
_SCRIPT_DIR = Path(__file__).parent
_BACKEND_DIR = _SCRIPT_DIR.parent / "backend"

try:
    from dotenv import load_dotenv
    _env_file = _BACKEND_DIR / ".env"
    if _env_file.exists():
        load_dotenv(_env_file)
        print(f"[env] Loaded {_env_file}")
    else:
        load_dotenv()  # try cwd/.env as fallback
except ImportError:
    pass  # dotenv not installed; rely on real env vars

# ---------------------------------------------------------------------------
# Validate required env vars early
# ---------------------------------------------------------------------------
MONGO_URL = os.environ.get("MONGO_URL", "")
DB_NAME = os.environ.get("DB_NAME", "jwoodtech")
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")

if not MONGO_URL:
    print(
        "\n[ERROR] MONGO_URL environment variable is not set.\n"
        "        Set it in your shell or create backend/.env with:\n"
        "          MONGO_URL=mongodb+srv://...\n"
        "          DB_NAME=jwoodtech\n"
    )
    sys.exit(1)

# ---------------------------------------------------------------------------
# Third-party imports (must be installed — see backend/requirements.txt)
# ---------------------------------------------------------------------------
try:
    import fitz  # PyMuPDF
    import httpx
    from bs4 import BeautifulSoup
    from motor.motor_asyncio import AsyncIOMotorClient
    from openpyxl import load_workbook
    from pydantic import BaseModel, ConfigDict, Field
except ImportError as e:
    print(
        f"\n[ERROR] Missing dependency: {e}\n"
        f"        Install with:  pip install -r {_BACKEND_DIR}/requirements.txt\n"
    )
    sys.exit(1)

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("crawl_vineyard")

# ---------------------------------------------------------------------------
# MongoDB client  (module-level, shared by all crawlers)
# ---------------------------------------------------------------------------
_mongo_client = AsyncIOMotorClient(MONGO_URL)
db = _mongo_client[DB_NAME]

# ---------------------------------------------------------------------------
# Constants — kept in sync with server.py
# ---------------------------------------------------------------------------
USER_AGENT = (
    "Mozilla/5.0 (compatible; JwoodVineyardBot/1.0; "
    "+https://jwoodtechnologies.com/)"
)
MAX_PAGES_PER_SOURCE = int(os.environ.get("MAX_PAGES_PER_SOURCE", "5000"))
MAX_DEPTH = int(os.environ.get("MAX_CRAWL_DEPTH", "10"))
REQUEST_TIMEOUT = 30.0
CRAWL_TIMEOUT_SECONDS = int(os.environ.get("CRAWL_TIMEOUT_SECONDS", "3600"))

MCO_TYPES = [
    ("ordinances", "Ordinances"),
    ("districts", "Districts"),
    ("resolutions", "Resolutions"),
    ("plan", "General Plan"),
    ("subdivords", "Subdivision Ordinances"),
    ("orddoc", "Ordinance Documents"),
    ("minutes", "Minutes"),
    ("zoning", "Zoning"),
    ("landscaping", "Landscaping"),
]

SECTION_PATTERNS = [
    re.compile(r"\b(?:Ordinance|Ord\.)\s*(?:No\.?\s*)?[\w\-\.]+", re.IGNORECASE),
    re.compile(r"\b(?:Resolution|Res\.)\s*(?:No\.?\s*)?[\w\-\.]+", re.IGNORECASE),
    re.compile(r"\bSection\s+\d+(?:[\-\.]\d+)*", re.IGNORECASE),
    re.compile(r"\bChapter\s+\d+(?:[\-\.]\d+)*", re.IGNORECASE),
    re.compile(r"\bTitle\s+\d+(?:[\-\.]\d+)*", re.IGNORECASE),
]

_DATE_PATS = [
    (re.compile(r"(\d{4})-(\d{1,2})-(\d{1,2})"), "ymd"),
    (re.compile(r"(\d{1,2})[/_-](\d{1,2})[/_-](\d{4})"), "mdy"),
    (
        re.compile(
            r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|"
            r"jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|"
            r"nov(?:ember)?|dec(?:ember)?)\s+(\d{1,2})(?:st|nd|rd|th)?[,]?\s+(\d{4})",
            re.IGNORECASE,
        ),
        "mname",
    ),
]
_MONTH_NAMES = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

MCO_SECTION_SPLIT = re.compile(
    r"<div\s+class=['\"]phx-name\s*['\"][^>]*>",
    re.IGNORECASE,
)

# Default sources — mirrors DEFAULT_VINEYARD_SOURCES in server.py
DEFAULT_VINEYARD_SOURCES = [
    ("https://vineyard.municipalcodeonline.com/", "Municipal Code"),
    ("https://www.vineyardutah.gov/", "Vineyard Utah (Official)"),
    ("https://vineyardut.portal.civicclerk.com/", "CivicClerk (Meetings)"),
    (
        "https://cms3.revize.com/revize/vineyard/Departmnts/"
        "Vineyard%20RDA%20past%20meetings%20index.xlsx",
        "RDA Past Meetings Index",
    ),
]


# ---------------------------------------------------------------------------
# Pydantic models (subset needed by crawlers)
# ---------------------------------------------------------------------------
class SourceRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    url: str
    label: str = ""
    status: str = "idle"
    pages_indexed: int = 0
    last_error: str = ""
    last_crawled_at: Optional[str] = None
    discovered: int = 0
    indexed: int = 0
    sections_indexed: int = 0
    pdfs_indexed: int = 0
    failed: int = 0
    max_depth_reached: int = 0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class DocumentRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    source_id: str
    source_root: str
    source_label: str = ""
    source_site: str = "other"
    url: str
    pdf_url: Optional[str] = None
    title: str
    section_ref: Optional[str] = None
    excerpt: str = ""
    content: str = ""
    is_pdf: bool = False
    is_xlsx: bool = False
    doc_type: str = "page"
    meeting_date: Optional[str] = None
    depth: int = 0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


# ---------------------------------------------------------------------------
# Pure helper functions — exact copies from server.py
# ---------------------------------------------------------------------------

def find_section_ref(text: str) -> Optional[str]:
    for pat in SECTION_PATTERNS:
        m = pat.search(text)
        if m:
            return m.group(0).strip()
    return None


def chunk_text(text: str, size: int = 1400, overlap: int = 200) -> List[str]:
    text = re.sub(r"\s+\n", "\n", text).strip()
    if not text:
        return []
    if len(text) <= size:
        return [text]
    chunks: List[str] = []
    start = 0
    while start < len(text):
        end = min(len(text), start + size)
        chunk = text[start:end]
        if end < len(text):
            last_break = max(chunk.rfind("\n\n"), chunk.rfind(". "))
            if last_break > size * 0.5:
                end = start + last_break + 1
                chunk = text[start:end]
        chunks.append(chunk.strip())
        start = end - overlap
        if start < 0:
            start = 0
        if end >= len(text):
            break
    return [c for c in chunks if c]


def html_to_text(html: str) -> tuple[str, str]:
    soup = BeautifulSoup(html, "lxml")
    title_tag = soup.find("title")
    title = title_tag.get_text(strip=True) if title_tag else ""
    if not title:
        h1 = soup.find("h1")
        if h1:
            title = h1.get_text(strip=True)
    for tag in soup(["script", "style", "noscript", "nav", "header", "footer", "aside"]):
        tag.decompose()
    text = soup.get_text("\n", strip=True)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return title, text


def pdf_to_text(data: bytes) -> str:
    try:
        with fitz.open(stream=data, filetype="pdf") as doc:
            parts = [page.get_text("text") for page in doc]
        return "\n\n".join(parts).strip()
    except Exception as exc:
        logger.warning("PDF parse failed: %s", exc)
        return ""


def xlsx_to_text(data: bytes) -> tuple[str, str, list[str]]:
    """Return (title, full_text, ordered_hyperlink_urls)."""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
        parts: list[str] = []
        title = wb.sheetnames[0] if wb.sheetnames else "Spreadsheet"
        seen_urls: set[str] = set()
        ordered_urls: list[str] = []
        for sheet in wb.worksheets:
            parts.append(f"## Sheet: {sheet.title}")
            for row in sheet.iter_rows():
                cells: list[str] = []
                for cell in row:
                    val = cell.value
                    href = (
                        cell.hyperlink.target
                        if cell.hyperlink and cell.hyperlink.target
                        else None
                    )
                    if val is None and not href:
                        continue
                    text = str(val).strip() if val is not None else ""
                    if href:
                        if href not in seen_urls:
                            seen_urls.add(href)
                            ordered_urls.append(href)
                        if text and text != href:
                            cells.append(f"{text} [{href}]")
                        else:
                            cells.append(href)
                    elif text:
                        cells.append(text)
                if cells:
                    parts.append(" | ".join(cells))
        return title, "\n".join(parts).strip(), ordered_urls
    except Exception as exc:
        logger.warning("XLSX parse failed: %s", exc)
        return "", "", []


def same_host(a: str, b: str) -> bool:
    try:
        return (
            urlparse(a).netloc.replace("www.", "")
            == urlparse(b).netloc.replace("www.", "")
        )
    except Exception:
        return False


def normalise(url: str) -> str:
    url, _ = urldefrag(url)
    return url.rstrip("/")


def _derive_source_site(source_root: str, url: str) -> str:
    h = (url or source_root or "").lower()
    if "civicclerk" in h:
        return "civicclerk"
    if "municipalcodeonline" in h or "municode" in h:
        return "municode"
    if "vineyardutah.gov" in h or "vineyardutah.org" in h:
        return "vineyardutah"
    if "transparent.utah.gov" in h:
        return "transparent"
    if "utah.gov" in h:
        return "utahgov"
    if "rda" in h or "revize.com" in h:
        return "rda"
    return "other"


def _derive_meeting_date(title: str, url: str) -> Optional[str]:
    for haystack in (title or "", url or ""):
        for pat, kind in _DATE_PATS:
            m = pat.search(haystack)
            if not m:
                continue
            try:
                if kind == "ymd":
                    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                elif kind == "mdy":
                    mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                else:  # mname
                    mo = _MONTH_NAMES[m.group(1)[:3].lower()]
                    d = int(m.group(2))
                    y = int(m.group(3))
                if 1990 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
                    return f"{y:04d}-{mo:02d}-{d:02d}"
            except Exception:
                continue
    return None


def _derive_doc_type(
    *, source_site: str, url: str, title: str, is_pdf: bool, is_xlsx: bool
) -> str:
    u = (url or "").lower()
    t = (title or "").lower()
    blob = f"{t} {u}"

    if is_xlsx:
        return "xlsx"
    if "resolution" in blob:
        return "resolution"
    if "ordinance" in blob or " ord " in f" {blob} ":
        return "ordinance"
    if "minutes" in blob:
        return "minutes"
    if "agenda packet" in blob or "agenda" in blob:
        return "agenda"
    if source_site == "rda":
        return "rda"
    if source_site == "transparent":
        return "transparency"
    if source_site == "municode":
        return "ordinance"
    if source_site == "civicclerk":
        if is_pdf:
            return "attachment"
        return "agenda"
    if is_pdf:
        return "pdf"
    return "page"


def _mco_parse_text_blob(text_html: str, parent_nameid: str) -> list[dict]:
    """Split MCO Text HTML into granular sections (one per phx-name marker)."""
    if not text_html:
        return []
    text_html = re.sub(r"<br\s*/?>", "\n", text_html, flags=re.IGNORECASE)
    parts = MCO_SECTION_SPLIT.split(text_html)
    out: list[dict] = []
    for piece in parts:
        if not piece.strip():
            continue
        soup = BeautifulSoup(piece, "lxml")
        nameid = parent_nameid
        first_link = soup.find("a", href=re.compile(r"#name="))
        if first_link and first_link.get("href"):
            nameid = first_link["href"].split("#name=", 1)[-1].split("&")[0]
        heading_el = soup.find(["h1", "h2", "h3", "h4"]) or first_link
        heading = (heading_el.get_text(strip=True) if heading_el else "").strip()
        text = soup.get_text("\n", strip=True)
        text = re.sub(r"\n{3,}", "\n\n", text)
        if len(text) < 40:
            continue
        out.append({"heading": heading or nameid, "nameid": nameid, "text": text})
    return out


# ---------------------------------------------------------------------------
# Async DB helpers
# ---------------------------------------------------------------------------

async def _store_chunks(
    chunks: list[str],
    *,
    source_id: str,
    source_root: str,
    source_label: str,
    url: str,
    title: str,
    pdf_url: Optional[str],
    is_pdf: bool,
    is_xlsx: bool,
    depth: int,
    crawl_run_id: Optional[str] = None,
    build_version: str = "v1",
) -> int:
    """Insert text chunks as DocumentRecord rows, tagged with build_version."""
    if not chunks:
        return 0
    site_bucket = _derive_source_site(source_root, url)
    meeting_dt = _derive_meeting_date(title, url)
    dtype = _derive_doc_type(
        source_site=site_bucket,
        url=url,
        title=title,
        is_pdf=is_pdf,
        is_xlsx=is_xlsx,
    )
    docs = []
    for chunk in chunks:
        rec = DocumentRecord(
            source_id=source_id,
            source_root=source_root,
            source_label=source_label,
            source_site=site_bucket,
            url=url,
            pdf_url=pdf_url,
            title=title,
            section_ref=find_section_ref(chunk[:2000]),
            excerpt=chunk[:220].strip(),
            content=chunk,
            is_pdf=is_pdf,
            is_xlsx=is_xlsx,
            doc_type=dtype,
            meeting_date=meeting_dt,
            depth=depth,
        )
        d = rec.model_dump()
        d["index_version"] = build_version
        if crawl_run_id:
            d["crawl_run_id"] = crawl_run_id
        docs.append(d)
    await db.documents.insert_many(docs)
    return len(docs)


async def _get_active_index() -> dict:
    meta = await db.index_meta.find_one({"kind": "active"}, {"_id": 0})
    if meta:
        return meta
    existing = await db.documents.count_documents({})
    if existing > 0:
        await db.documents.update_many(
            {"index_version": {"$exists": False}},
            {"$set": {"index_version": "v1"}},
        )
        meta = {
            "kind": "active",
            "version": "v1",
            "doc_count": existing,
            "locked_at": datetime.now(timezone.utc).isoformat(),
            "locked_by": "bootstrap",
        }
        await db.index_meta.update_one({"kind": "active"}, {"$set": meta}, upsert=True)
        return meta
    return {"kind": "active", "version": None, "doc_count": 0}


async def lock_index_version(version: str, doc_count: int, locked_by: str = "script") -> None:
    """Atomically promote `version` to be the active archive."""
    # Archive the previous active version as prev-<old>
    cur = await _get_active_index()
    prev_version = cur.get("version")
    if prev_version and prev_version != version:
        prev_renamed = f"prev-{prev_version}"
        # Purge older prev-* snapshots (keep at most one rollback copy)
        await db.documents.delete_many({"index_version": {"$regex": "^prev-"}})
        await db.documents.update_many(
            {"index_version": prev_version},
            {"$set": {"index_version": prev_renamed}},
        )
        logger.info("Previous version %s archived as %s", prev_version, prev_renamed)

    await db.index_meta.update_one(
        {"kind": "active"},
        {
            "$set": {
                "version": version,
                "doc_count": int(doc_count),
                "locked_at": datetime.now(timezone.utc).isoformat(),
                "locked_by": locked_by,
            }
        },
        upsert=True,
    )
    logger.info("Index locked: version=%s doc_count=%d by=%s", version, doc_count, locked_by)


# ---------------------------------------------------------------------------
# Source seeding
# ---------------------------------------------------------------------------

async def seed_sources() -> list[dict]:
    """Ensure DEFAULT_VINEYARD_SOURCES exist in db.sources. Return all sources."""
    for url, label in DEFAULT_VINEYARD_SOURCES:
        existing = await db.sources.find_one({"url": url}, {"_id": 0})
        if existing:
            if not existing.get("label"):
                await db.sources.update_one(
                    {"id": existing["id"]}, {"$set": {"label": label}}
                )
            continue
        rec = SourceRecord(url=url, label=label)
        await db.sources.insert_one(rec.model_dump())
        logger.info("Seeded source: %s (%s)", url, label)
    return await db.sources.find({}, {"_id": 0}).sort("created_at", 1).to_list(50)


# ---------------------------------------------------------------------------
# Crawler: CivicClerk OData REST API
# ---------------------------------------------------------------------------

async def _crawl_civicclerk(
    source_id: str,
    root_url: str,
    source_label: str = "",
    crawl_run_id: Optional[str] = None,
    build_version: str = "v1",
) -> dict:
    """Deep CivicClerk crawl using the public OData REST API.

    1. Pages through /v1/Events (both directions) to enumerate all meetings.
    2. Fetches /v1/Meetings/{agendaId} for each event.
    3. Indexes the agenda outline as a searchable record.
    4. Downloads every published file via plainText=true; falls back to PDF.
    """
    parsed = urlparse(root_url)
    portal_host = parsed.netloc
    api_host = portal_host.replace(".portal.", ".api.")
    if not api_host.endswith(".civicclerk.com"):
        api_host = "vineyardut.api.civicclerk.com"
    api_base = f"https://{api_host}/v1"

    discovered = 0
    indexed = 0
    errors = 0

    headers = {"User-Agent": USER_AGENT, "Accept": "application/json"}
    async with httpx.AsyncClient(
        timeout=REQUEST_TIMEOUT,
        follow_redirects=True,
        headers=headers,
    ) as http:
        # Enumerate every event by paginating both asc and desc
        events: list[dict] = []
        for direction in ("desc", "asc"):
            skip = 0
            empty_streak = 0
            while True:
                params = {
                    "$orderby": f"startDateTime {direction}",
                    "$top": "100",
                    "$skip": str(skip),
                }
                try:
                    resp = await http.get(f"{api_base}/Events", params=params)
                    resp.raise_for_status()
                    page = resp.json().get("value", [])
                except Exception as exc:
                    logger.warning(
                        "civicclerk events page failed (%s skip=%d): %s", direction, skip, exc
                    )
                    errors += 1
                    break
                if not page:
                    break
                events.extend(page)
                advance = max(len(page), 15)
                skip += advance
                if skip > 5000:
                    break
                if len(page) < 5:
                    empty_streak += 1
                    if empty_streak >= 2:
                        break
                else:
                    empty_streak = 0

        # Dedupe by event id
        seen_ids: set[int] = set()
        unique_events: list[dict] = []
        for ev in events:
            eid = ev.get("id")
            if eid is None or eid in seen_ids:
                continue
            seen_ids.add(eid)
            unique_events.append(ev)
        discovered = len(unique_events)
        logger.info("civicclerk: %d unique events for %s", discovered, api_base)

        for ev in unique_events:
            event_id = ev.get("id")
            agenda_id = ev.get("agendaId")
            event_name = (ev.get("eventName") or "Meeting").strip()
            event_date = (ev.get("startDateTime") or "")[:10]
            category = (ev.get("categoryName") or "").strip()
            portal_url = f"https://{portal_host}/event/{event_id}/files"

            published_files: list[dict] = []
            outline_text = ""
            if agenda_id:
                try:
                    mr = await http.get(f"{api_base}/Meetings/{agenda_id}")
                    if mr.status_code == 200:
                        meeting = mr.json()
                        published_files = meeting.get("publishedFiles") or []
                        bits: list[str] = []
                        for it in meeting.get("items") or []:
                            num = (it.get("agendaObjectItemOutlineNumber") or "").strip()
                            name = (it.get("agendaObjectItemName") or "").strip()
                            desc = (it.get("agendaObjectItemDescription") or "").strip()
                            if name:
                                bits.append(f"{num} {name}".strip())
                            if desc:
                                bits.append(desc)
                        outline_text = "\n".join(bits).strip()
                except Exception as exc:
                    logger.warning(
                        "civicclerk meeting fetch failed event=%s agenda=%s: %s",
                        event_id, agenda_id, exc,
                    )
                    errors += 1

            # Index the outline record for this meeting
            meta_lines = [event_name]
            if event_date:
                meta_lines.append(f"Meeting Date: {event_date}")
            if category:
                meta_lines.append(f"Category: {category}")
            full_outline = "\n".join(meta_lines)
            if outline_text:
                full_outline = f"{full_outline}\n\nAGENDA OUTLINE\n{outline_text}"
            if len(full_outline) > 200:
                chunks = chunk_text(full_outline)
                added = await _store_chunks(
                    chunks,
                    source_id=source_id,
                    source_root=root_url,
                    source_label=source_label,
                    url=portal_url,
                    title=f"{event_name} — {event_date}".strip(" —"),
                    pdf_url=None,
                    is_pdf=False,
                    is_xlsx=False,
                    depth=1,
                    crawl_run_id=crawl_run_id,
                    build_version=build_version,
                )
                if added:
                    indexed += 1

            # Index each published file
            for pf in published_files:
                file_id = pf.get("fileId")
                if not file_id:
                    continue
                ftype = (pf.get("type") or "File").strip()
                fname = (pf.get("name") or f"{ftype} {file_id}").strip()
                stream_pdf = (
                    f"{api_base}/Meetings/GetMeetingFileStream"
                    f"(fileId={file_id},plainText=false)"
                )
                stream_txt = (
                    f"{api_base}/Meetings/GetMeetingFileStream"
                    f"(fileId={file_id},plainText=true)"
                )
                try:
                    fr = await http.get(stream_txt)
                except Exception as exc:
                    logger.warning(
                        "civicclerk file fetch failed event=%s file=%s: %s",
                        event_id, file_id, exc,
                    )
                    errors += 1
                    continue
                text = ""
                if fr.status_code == 200:
                    text = (fr.text or "").strip()
                if len(text) < 80:
                    try:
                        pr = await http.get(stream_pdf)
                        if (
                            pr.status_code == 200
                            and "pdf" in (pr.headers.get("content-type", "")).lower()
                        ):
                            text = (pdf_to_text(pr.content) or "").strip()
                    except Exception as exc:
                        logger.warning(
                            "civicclerk pdf fallback failed event=%s file=%s: %s",
                            event_id, file_id, exc,
                        )
                if len(text) < 80:
                    continue
                header = f"{event_name} — {event_date} — {ftype}: {fname}"
                body = f"{header}\n\n{text}"
                chunks = chunk_text(body)
                added = await _store_chunks(
                    chunks,
                    source_id=source_id,
                    source_root=root_url,
                    source_label=source_label,
                    url=portal_url,
                    title=f"{event_name} — {ftype} — {fname}".strip(),
                    pdf_url=stream_pdf,
                    is_pdf=True,
                    is_xlsx=False,
                    depth=2,
                    crawl_run_id=crawl_run_id,
                    build_version=build_version,
                )
                if added:
                    indexed += 1

    await db.sources.update_one(
        {"id": source_id},
        {"$set": {"discovered": discovered, "max_depth_reached": 2}},
    )
    logger.info(
        "civicclerk done %s: events=%d indexed=%d errors=%d",
        root_url, discovered, indexed, errors,
    )
    return {"discovered": discovered, "indexed": indexed, "errors": errors}


# ---------------------------------------------------------------------------
# Crawler: Generic BFS (HTML + PDF + XLSX) — used for vineyardutah.gov etc.
# ---------------------------------------------------------------------------

async def _crawl_generic(
    source_id: str,
    root_url: str,
    source_label: str = "",
    crawl_run_id: Optional[str] = None,
    build_version: str = "v1",
) -> dict:
    """Recursive polite crawl within host. Handles HTML, PDF, XLSX."""
    seen: set[str] = set()
    queue: list[tuple[str, int]] = [(root_url, 0)]
    pages_indexed = 0
    errors = 0

    headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    async with httpx.AsyncClient(
        timeout=REQUEST_TIMEOUT,
        follow_redirects=True,
        headers=headers,
    ) as http:
        while queue and pages_indexed < MAX_PAGES_PER_SOURCE:
            url, depth = queue.pop(0)
            url = normalise(url)
            if url in seen:
                continue
            seen.add(url)

            try:
                resp = await http.get(url)
            except Exception as exc:
                logger.warning("fetch failed %s: %s", url, exc)
                errors += 1
                continue
            if resp.status_code >= 400:
                continue

            ctype = resp.headers.get("content-type", "").lower()
            low = url.lower()
            is_pdf = "application/pdf" in ctype or low.endswith(".pdf")
            is_xlsx = (
                "spreadsheetml" in ctype
                or "ms-excel" in ctype
                or low.endswith(".xlsx")
                or low.endswith(".xls")
            )

            if is_pdf:
                text = pdf_to_text(resp.content)
                if text:
                    title = url.rsplit("/", 1)[-1] or "PDF Document"
                    chunks = chunk_text(text)
                    added = await _store_chunks(
                        chunks,
                        source_id=source_id,
                        source_root=root_url,
                        source_label=source_label,
                        url=url,
                        title=title,
                        pdf_url=url,
                        is_pdf=True,
                        is_xlsx=False,
                        depth=depth,
                        crawl_run_id=crawl_run_id,
                        build_version=build_version,
                    )
                    if added:
                        pages_indexed += 1
                continue

            if is_xlsx:
                title, text, xlsx_links = xlsx_to_text(resp.content)
                if text:
                    title = title or url.rsplit("/", 1)[-1] or "Spreadsheet"
                    chunks = chunk_text(text, size=1800, overlap=200)
                    added = await _store_chunks(
                        chunks,
                        source_id=source_id,
                        source_root=root_url,
                        source_label=source_label,
                        url=url,
                        title=title,
                        pdf_url=None,
                        is_pdf=False,
                        is_xlsx=True,
                        depth=depth,
                        crawl_run_id=crawl_run_id,
                        build_version=build_version,
                    )
                    if added:
                        pages_indexed += 1
                if xlsx_links:
                    active_meta = await _get_active_index()
                    active_v = active_meta.get("version") or "v1"
                    for link in xlsx_links:
                        norm = normalise(link)
                        if norm in seen:
                            continue
                        if "civicclerk.com" in norm:
                            continue
                        already = await db.documents.find_one(
                            {"index_version": active_v, "url": norm}, {"_id": 0, "id": 1}
                        )
                        if already:
                            continue
                        queue.append((norm, depth + 1))
                continue

            if "text/html" not in ctype and "xml" not in ctype:
                continue

            title, text = html_to_text(resp.text)
            if text and len(text) > 80:
                chunks = chunk_text(text)
                soup_ctx = BeautifulSoup(resp.text, "lxml")
                pdf_link = None
                for a in soup_ctx.find_all("a", href=True):
                    href = a["href"]
                    if href.lower().endswith(".pdf"):
                        pdf_link = urljoin(url, href)
                        break
                added = await _store_chunks(
                    chunks,
                    source_id=source_id,
                    source_root=root_url,
                    source_label=source_label,
                    url=url,
                    title=title or url,
                    pdf_url=pdf_link,
                    is_pdf=False,
                    is_xlsx=False,
                    depth=depth,
                    crawl_run_id=crawl_run_id,
                    build_version=build_version,
                )
                if added:
                    pages_indexed += 1

            if depth < MAX_DEPTH:
                soup = BeautifulSoup(resp.text, "lxml")
                for a in soup.find_all("a", href=True):
                    href = a["href"].strip()
                    if href.startswith(("mailto:", "tel:", "javascript:")):
                        continue
                    full = normalise(urljoin(url, href))
                    if not full.startswith("http"):
                        continue
                    low2 = full.lower()
                    is_file = low2.endswith((".pdf", ".xlsx", ".xls"))
                    if not is_file and not same_host(full, root_url):
                        continue
                    if full in seen:
                        continue
                    queue.append((full, depth + 1))

            await asyncio.sleep(0.1)

    pdf_count = await db.documents.count_documents(
        {"source_id": source_id, "crawl_run_id": crawl_run_id, "is_pdf": True}
    ) if crawl_run_id else 0
    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "discovered": len(seen),
                "pdfs_indexed": pdf_count,
                "max_depth_reached": MAX_DEPTH,
                "failed": max(0, len(seen) - pages_indexed),
            }
        },
    )
    logger.info("generic crawl done %s: pages=%d errors=%d", root_url, pages_indexed, errors)
    return {"discovered": len(seen), "indexed": pages_indexed, "errors": errors}


# ---------------------------------------------------------------------------
# Crawler: MunicipalCodeOnline (Playwright)
# ---------------------------------------------------------------------------

async def _crawl_mco(
    source_id: str,
    root_url: str,
    source_label: str = "",
    crawl_run_id: Optional[str] = None,
    build_version: str = "v1",
) -> dict:
    """Deep crawl of municipalcodeonline.com using Playwright + click-driven XHRs."""
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        logger.error(
            "Playwright is not installed. Run: pip install playwright && playwright install chromium"
        )
        return {"discovered": 0, "indexed": 0, "errors": 1}

    discovered = 0
    indexed = 0
    failed = 0
    total_stored = 0
    max_depth = 0

    parsed = urlparse(root_url)
    origin = f"{parsed.scheme}://{parsed.netloc}"

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        ctx = await browser.new_context(user_agent=USER_AGENT)
        page = await ctx.new_page()

        for type_key, type_label in MCO_TYPES:
            book_url = f"{origin}/book?type={type_key}"
            expand_body_holder: dict[str, str] = {}
            content_bodies: dict[str, str] = {}

            async def on_resp(resp, _tk=type_key):
                url = resp.url
                try:
                    if "/book/expand" in url and resp.status == 200:
                        expand_body_holder["body"] = await resp.text()
                    elif "/book/content" in url and resp.status == 200:
                        q = parse_qs(urlparse(url).query)
                        nm = q.get("name", ["?"])[0]
                        t = q.get("type", ["?"])[0]
                        if t == _tk and nm not in content_bodies:
                            content_bodies[nm] = await resp.text()
                except Exception:
                    pass

            page.on("response", on_resp)
            try:
                await page.goto(book_url, wait_until="networkidle", timeout=60000)
                await page.wait_for_timeout(2500)
            except Exception as exc:
                failed += 1
                logger.warning("mco goto %s failed: %s", book_url, exc)
                page.remove_listener("response", on_resp)
                continue

            tree: list[dict] = []
            try:
                tree = json.loads(expand_body_holder.get("body", "[]"))
            except Exception:
                tree = []
            discovered += len(tree)

            for node in tree:
                nm = node.get("NameId") or node.get("name", "")
                label = node.get("name", nm)
                if not label:
                    continue
                if nm in content_bodies:
                    continue
                try:
                    await page.click(f'text="{label}"', timeout=4000)
                    await page.wait_for_timeout(900)
                except Exception:
                    failed += 1
                    continue

            page.remove_listener("response", on_resp)

            for nm, raw_json in content_bodies.items():
                try:
                    j = json.loads(raw_json)
                except Exception:
                    failed += 1
                    continue
                text_html = j.get("Text") or j.get("text") or ""
                if not text_html:
                    continue
                sections = _mco_parse_text_blob(text_html, nm)
                if not sections:
                    soup = BeautifulSoup(text_html, "lxml")
                    raw = soup.get_text("\n", strip=True)
                    chunks = chunk_text(raw)
                    url = f"{origin}/book?type={type_key}#name={nm}"
                    added = await _store_chunks(
                        chunks,
                        source_id=source_id,
                        source_root=root_url,
                        source_label=source_label or type_label,
                        url=url,
                        title=f"{type_label} · {nm.replace('_', ' ')}",
                        pdf_url=None,
                        is_pdf=False,
                        is_xlsx=False,
                        depth=1,
                        crawl_run_id=crawl_run_id,
                        build_version=build_version,
                    )
                    total_stored += added
                    indexed += 1 if added else 0
                    continue

                for sec in sections:
                    sec_nameid = sec["nameid"]
                    url = f"{origin}/book?type={type_key}#name={sec_nameid}"
                    title = sec["heading"] if sec["heading"] else sec_nameid
                    text = sec["text"]
                    chunks = chunk_text(text, size=1400, overlap=200)
                    added = await _store_chunks(
                        chunks,
                        source_id=source_id,
                        source_root=root_url,
                        source_label=source_label or type_label,
                        url=url,
                        title=f"{type_label} · {title}",
                        pdf_url=None,
                        is_pdf=False,
                        is_xlsx=False,
                        depth=2,
                        crawl_run_id=crawl_run_id,
                        build_version=build_version,
                    )
                    total_stored += added
                    if added:
                        indexed += 1
                        max_depth = max(max_depth, 2)

        await browser.close()

    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "discovered": discovered,
                "failed": failed,
                "max_depth_reached": max_depth or 2,
            }
        },
    )
    logger.info(
        "mco done %s: discovered=%d indexed=%d stored=%d failed=%d",
        root_url, discovered, indexed, total_stored, failed,
    )
    return {"discovered": discovered, "indexed": indexed, "errors": failed}


# ---------------------------------------------------------------------------
# Dispatcher: picks crawler by URL host, mirrors server.py crawl_source_task
# ---------------------------------------------------------------------------

async def crawl_source(
    source_id: str,
    root_url: str,
    source_label: str = "",
    build_version: str = "v1",
) -> dict:
    """Run the appropriate crawler for root_url. Returns result dict."""
    host = urlparse(root_url).netloc.lower()
    run_id = uuid.uuid4().hex

    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "status": "crawling",
                "last_error": "",
                "current_run_id": run_id,
                "current_run_started_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )

    async def _dispatch():
        if "municipalcodeonline.com" in host:
            return await _crawl_mco(source_id, root_url, source_label, run_id, build_version)
        elif "civicclerk.com" in host:
            return await _crawl_civicclerk(source_id, root_url, source_label, run_id, build_version)
        else:
            return await _crawl_generic(source_id, root_url, source_label, run_id, build_version)

    try:
        result = await asyncio.wait_for(_dispatch(), timeout=CRAWL_TIMEOUT_SECONDS)
    except asyncio.TimeoutError:
        logger.warning("crawl timeout (%ds) for %s", CRAWL_TIMEOUT_SECONDS, root_url)
        await db.documents.delete_many({"source_id": source_id, "crawl_run_id": run_id})
        await db.sources.update_one(
            {"id": source_id},
            {
                "$set": {
                    "status": "timeout",
                    "last_error": f"timeout after {CRAWL_TIMEOUT_SECONDS}s",
                    "last_finished_at": datetime.now(timezone.utc).isoformat(),
                },
                "$unset": {"current_run_id": "", "current_run_started_at": ""},
            },
        )
        return {"discovered": 0, "indexed": 0, "errors": 1, "status": "timeout"}
    except Exception as exc:
        logger.exception("dispatcher error for %s: %s", root_url, exc)
        await db.documents.delete_many({"source_id": source_id, "crawl_run_id": run_id})
        await db.sources.update_one(
            {"id": source_id},
            {
                "$set": {
                    "status": "error",
                    "last_error": str(exc)[:400],
                    "last_finished_at": datetime.now(timezone.utc).isoformat(),
                },
                "$unset": {"current_run_id": "", "current_run_started_at": ""},
            },
        )
        return {"discovered": 0, "indexed": 0, "errors": 1, "status": "error"}

    # Success: strip the transient crawl_run_id tag so these docs are simply
    # tagged with build_version. Also persist final stats.
    new_count = await db.documents.count_documents(
        {"source_id": source_id, "crawl_run_id": run_id}
    )
    if new_count == 0:
        await db.sources.update_one(
            {"id": source_id},
            {
                "$set": {
                    "status": "error",
                    "last_error": "no documents found",
                    "last_finished_at": datetime.now(timezone.utc).isoformat(),
                },
                "$unset": {"current_run_id": "", "current_run_started_at": ""},
            },
        )
        result["status"] = "error"
        return result

    await db.documents.update_many(
        {"source_id": source_id, "crawl_run_id": run_id},
        {"$unset": {"crawl_run_id": ""}},
    )
    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "status": "done",
                "last_error": "",
                "last_crawled_at": datetime.now(timezone.utc).isoformat(),
                "last_finished_at": datetime.now(timezone.utc).isoformat(),
                "pages_indexed": new_count,
            },
            "$unset": {"current_run_id": "", "current_run_started_at": ""},
        },
    )
    result["status"] = "done"
    result["doc_count"] = new_count
    return result


# ---------------------------------------------------------------------------
# CLI commands
# ---------------------------------------------------------------------------

async def cmd_status() -> None:
    """Print current index and source status."""
    meta = await _get_active_index()
    total = await db.documents.count_documents({})
    sources = await db.sources.find({}, {"_id": 0}).sort("created_at", 1).to_list(50)

    print("\n=== Vineyard City Index Status ===")
    print(f"  MongoDB DB:      {DB_NAME}")
    print(f"  Active version:  {meta.get('version') or '(none)'}")
    print(f"  Active doc_count:{meta.get('doc_count', 0)}")
    print(f"  Locked at:       {meta.get('locked_at') or '(never)'}")
    print(f"  Total docs in DB:{total}")
    print(f"\n  Sources ({len(sources)} total):")
    for s in sources:
        sid_count = await db.documents.count_documents({"source_id": s["id"]})
        print(
            f"    [{s.get('status', 'idle'):8s}] {s.get('label', 'Unlabeled'):30s}  "
            f"docs={sid_count:5d}  url={s.get('url', '')[:60]}"
        )

    versions_pipeline = [
        {"$group": {"_id": "$index_version", "n": {"$sum": 1}}},
        {"$sort": {"n": -1}},
    ]
    versions = await db.documents.aggregate(versions_pipeline).to_list(20)
    if versions:
        print("\n  Index versions:")
        for v in versions:
            tag = ""
            if v["_id"] == meta.get("version"):
                tag = " <-- ACTIVE"
            print(f"    {(v['_id'] or '(untagged)'):40s}  {v['n']:6d} docs{tag}")
    print()


async def cmd_dry_run() -> None:
    """Connect and report existing data without crawling."""
    print("\n[dry-run] Connecting to MongoDB Atlas...")
    await cmd_status()
    print("[dry-run] No crawl performed.")


async def cmd_crawl(
    source_filter: Optional[str] = None,
    lock: bool = True,
) -> None:
    """Seed sources, crawl, and optionally lock the new version."""
    build_ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    build_version = f"build-{build_ts}"
    logger.info("Build version: %s", build_version)

    sources = await seed_sources()

    # Filter by --source argument
    FILTER_MAP = {
        "civicclerk": "civicclerk",
        "municode": "municipalcodeonline",
        "vineyard": "vineyardutah.gov",
    }
    if source_filter:
        key = source_filter.lower()
        match_str = FILTER_MAP.get(key, key)
        sources = [s for s in sources if match_str in s.get("url", "").lower()]
        if not sources:
            print(
                f"[ERROR] No source matched --source {source_filter!r}. "
                f"Valid values: civicclerk, municode, vineyard"
            )
            return

    summary: list[dict] = []
    total_indexed = 0

    for s in sources:
        url = s.get("url", "")
        label = s.get("label", "")
        sid = s["id"]
        logger.info("Crawling: %s (%s)", label, url)
        try:
            result = await crawl_source(sid, url, label, build_version)
            result["url"] = url
            result["label"] = label
            summary.append(result)
            total_indexed += result.get("doc_count", result.get("indexed", 0))
        except Exception as exc:
            logger.exception("Crawl failed for %s: %s", url, exc)
            summary.append({"url": url, "label": label, "status": "error", "errors": 1})

    # Print summary
    print("\n=== Crawl Summary ===")
    print(f"  Build version: {build_version}")
    for r in summary:
        status = r.get("status", "?")
        n = r.get("doc_count", r.get("indexed", 0))
        err = r.get("errors", 0)
        print(
            f"  [{status:8s}] {r.get('label', ''):30s}  "
            f"docs={n:5d}  errors={err}  url={r.get('url', '')[:60]}"
        )
    print(f"\n  Total documents indexed into {build_version}: {total_indexed}")

    if lock and total_indexed > 0:
        # Count actual docs with this build_version (more accurate than sum)
        real_count = await db.documents.count_documents({"index_version": build_version})
        await lock_index_version(build_version, real_count, locked_by="crawl_vineyard.py")
        print(f"\n  Index locked: {build_version} ({real_count} docs)")
    elif lock and total_indexed == 0:
        print("\n  [WARN] No documents were indexed — skipping lock to preserve previous index.")
    else:
        print(
            f"\n  [INFO] --no-lock specified. To promote this build, run:\n"
            f"         python scripts/crawl_vineyard.py --status"
        )

    print()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="crawl_vineyard.py",
        description="Vineyard City government research tool — standalone crawl script",
    )
    group = p.add_mutually_exclusive_group()
    group.add_argument(
        "--dry-run",
        action="store_true",
        help="Connect to MongoDB and report existing data only (no crawl)",
    )
    group.add_argument(
        "--status",
        action="store_true",
        help="Print current index and source status",
    )
    p.add_argument(
        "--source",
        metavar="NAME",
        default=None,
        help=(
            "Crawl only a specific source. "
            "NAME must be one of: civicclerk, municode, vineyard"
        ),
    )
    p.add_argument(
        "--no-lock",
        action="store_true",
        help="Skip locking the new build version after crawling",
    )
    return p


async def _main() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    if args.status:
        await cmd_status()
    elif args.dry_run:
        await cmd_dry_run()
    else:
        await cmd_crawl(
            source_filter=args.source,
            lock=not args.no_lock,
        )

    _mongo_client.close()


if __name__ == "__main__":
    asyncio.run(_main())
