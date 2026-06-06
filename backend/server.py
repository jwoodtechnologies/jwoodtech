"""Jwood Technologies + Vineyard Scraper API."""
from __future__ import annotations

import asyncio
import collections
import io
import json
import logging
import math
import os
import re
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import List, Optional
from urllib.parse import urldefrag, urljoin, urlparse

import fitz  # PyMuPDF
import httpx
import resend
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.interval import IntervalTrigger
from bs4 import BeautifulSoup
from ddgs import DDGS
from dotenv import load_dotenv
from fastapi import APIRouter, BackgroundTasks, FastAPI, HTTPException
from motor.motor_asyncio import AsyncIOMotorClient
from openai import AsyncOpenAI
from openpyxl import load_workbook
from playwright.async_api import async_playwright
from pydantic import BaseModel, ConfigDict, EmailStr, Field
from rank_bm25 import BM25Okapi
from starlette.middleware.cors import CORSMiddleware

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("jwood")

# ---------------------------------------------------------------------------
# Config / clients
# ---------------------------------------------------------------------------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

VINEYARD_PASSWORD = os.environ.get("VINEYARD_PASSWORD", "555")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "7607")
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "")
RECIPIENT_EMAIL = os.environ.get("RECIPIENT_EMAIL", "")

EMBED_MODEL = "text-embedding-3-small"
SUMMARY_MODEL = "gpt-4o-mini"

openai_client: Optional[AsyncOpenAI] = (
    AsyncOpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None
)
if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
app = FastAPI(title="Jwood Technologies API")
api = APIRouter(prefix="/api")

SERVER_STARTED_AT = datetime.now(timezone.utc).isoformat()
CRAWL_INTERVAL_DAYS = int(os.environ.get("CRAWL_INTERVAL_DAYS", "7"))
CRAWL_TIMEOUT_SECONDS = int(os.environ.get("CRAWL_TIMEOUT_SECONDS", "1800"))  # 30 min default
scheduler = AsyncIOScheduler(timezone="UTC")


async def _metric_inc(key: str, n: int = 1) -> None:
    """Atomic counter bump stored in the `metrics` collection."""
    try:
        await db.metrics.update_one(
            {"_id": "global"}, {"$inc": {key: int(n)}}, upsert=True
        )
    except Exception as exc:
        logger.debug("metric inc failed (%s): %s", key, exc)


# ---------------------------------------------------------------------------
# Vineyard index — versioned, lockable archive
# ---------------------------------------------------------------------------
# Users only ever read the LOCKED archive they're queried against — even
# while a re-crawl is mutating the database. Single source of truth:
# `db.index_meta` doc with `kind == "active"`. Each chunk in `db.documents`
# carries `index_version`; user APIs filter by the active version. Crawls
# write into a fresh version so partial work is INVISIBLE until the admin
# atomically flips the pointer.


async def _get_active_index() -> dict:
    """Return the active-index meta doc, bootstrapping if first call."""
    meta = await db.index_meta.find_one({"kind": "active"}, {"_id": 0})
    if meta:
        return meta
    existing = await db.documents.count_documents({})
    if existing > 0:
        # Lock whatever's already in the DB as v1 so the production page
        # keeps showing the same archive without manual intervention.
        await db.documents.update_many(
            {"index_version": {"$exists": False}},
            {"$set": {"index_version": "v1"}},
        )
        meta = {
            "kind": "active",
            "version": "v1",
            "doc_count": existing,
            "locked_at": datetime.now(timezone.utc).isoformat(),
            "locked_by": "bootstrap",
        }
        await db.index_meta.update_one(
            {"kind": "active"}, {"$set": meta}, upsert=True
        )
        return meta
    return {"kind": "active", "version": None, "doc_count": 0}


async def get_active_filter() -> dict:
    """Mongo filter scoped to the locked active archive."""
    meta = await _get_active_index()
    v = meta.get("version")
    if not v:
        return {"index_version": "__none__"}
    return {"index_version": v}


async def lock_index_version(
    version: str, doc_count: int, locked_by: str = "admin"
):
    """Atomically promote `version` to be the active archive."""
    await db.index_meta.update_one(
        {"kind": "active"},
        {
            "$set": {
                "version": version,
                "doc_count": int(doc_count),
                "locked_at": datetime.now(timezone.utc).isoformat(),
                "locked_by": locked_by,
            }
        },
        upsert=True,
    )
    logger.info(
        "vineyard index locked: version=%s doc_count=%d by=%s",
        version,
        doc_count,
        locked_by,
    )


async def scheduled_refresh():
    """Fire-and-forget re-crawl of every configured source."""
    sources = await db.sources.find({}, {"_id": 0}).to_list(50)
    if not sources:
        return
    for s in sources:
        await db.sources.update_one(
            {"id": s["id"]},
            {"$set": {"status": "crawling", "last_error": ""}},
        )
        asyncio.create_task(
            crawl_source_task(s["id"], s["url"], s.get("label", ""))
        )
    logger.info(
        "scheduled weekly re-crawl triggered for %d sources", len(sources)
    )


def _next_scheduled_run_iso() -> Optional[str]:
    try:
        job = scheduler.get_job("weekly-refresh")
        if job and job.next_run_time:
            return job.next_run_time.astimezone(timezone.utc).isoformat()
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class ContactCreate(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = ""
    project_type: str
    description: str
    budget: str
    timeline: str


class ContactRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    phone: str = ""
    project_type: str
    description: str
    budget: str
    timeline: str
    email_sent: bool = False
    created_at: str = Field(
        default_factory=lambda: datetime.now(timezone.utc).isoformat()
    )


class AuthRequest(BaseModel):
    password: str


class SourceCreate(BaseModel):
    url: str
    label: Optional[str] = ""


class SourceRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    url: str
    label: str = ""
    status: str = "idle"
    pages_indexed: int = 0
    last_error: str = ""
    last_crawled_at: Optional[str] = None
    # Deep-crawl report metrics
    discovered: int = 0
    indexed: int = 0
    sections_indexed: int = 0
    pdfs_indexed: int = 0
    failed: int = 0
    max_depth_reached: int = 0
    created_at: str = Field(
        default_factory=lambda: datetime.now(timezone.utc).isoformat()
    )


class DocumentRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    source_id: str
    source_root: str
    source_label: str = ""
    # Filterable provenance — derived from the source URL host:
    #   civicclerk | municode | vineyardutah | rda | other
    source_site: str = "other"
    url: str
    pdf_url: Optional[str] = None
    title: str
    section_ref: Optional[str] = None
    excerpt: str = ""
    content: str = ""
    is_pdf: bool = False
    is_xlsx: bool = False
    # Filterable: page | pdf | agenda | minutes | ordinance | attachment | rda | xlsx
    doc_type: str = "page"
    # ISO date string when we can extract one from the URL or page title
    meeting_date: Optional[str] = None
    depth: int = 0
    created_at: str = Field(
        default_factory=lambda: datetime.now(timezone.utc).isoformat()
    )


class SearchRequest(BaseModel):
    query: str
    # Source filter — accepts EITHER a `source_site` bucket
    # (civicclerk|municode|vineyardutah|utahgov|transparent) for backward
    # compat, OR a literal `source_id` UUID. The chip row in the UI now
    # uses source_id so newly added sources auto-create working filters.
    site: Optional[str] = None
    source_id: Optional[str] = None
    doc_type: Optional[str] = None  # page | pdf | agenda | minutes | ordinance | attachment | rda | xlsx
    date_from: Optional[str] = None  # ISO yyyy-mm-dd
    date_to: Optional[str] = None


class Citation(BaseModel):
    title: str
    source_label: str
    source_site: str = "other"
    doc_type: str = "page"
    meeting_date: Optional[str] = None
    url: str
    pdf_url: Optional[str] = None
    section_ref: Optional[str] = None
    excerpt: str
    score: float


class SearchResponse(BaseModel):
    answer: str
    citations: List[Citation]
    has_results: bool
    query: str


# ---------------------------------------------------------------------------
# Root + Contact
# ---------------------------------------------------------------------------
@api.get("/")
async def root():
    return {"service": "jwood-technologies", "ok": True}


def _inquiry_html(r: ContactRecord) -> str:
    def esc(v: str) -> str:
        return (
            (v or "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    row = (
        "<tr><td style='padding:10px 14px;background:#f6f7f9;color:#6b7280;"
        "font-family:Arial,sans-serif;font-size:12px;letter-spacing:.08em;"
        "text-transform:uppercase;width:140px;vertical-align:top'>{k}</td>"
        "<td style='padding:10px 14px;color:#0b0f19;font-family:Arial,sans-serif;"
        "font-size:14px;vertical-align:top'>{v}</td></tr>"
    )
    rows = "".join(
        row.format(k=k, v=esc(v).replace("\n", "<br>"))
        for k, v in [
            ("Name", r.name),
            ("Email", r.email),
            ("Phone", r.phone or "—"),
            ("Project type", r.project_type),
            ("Budget", r.budget),
            ("Timeline", r.timeline),
            ("Description", r.description),
        ]
    )
    return (
        "<!doctype html><html><body style='margin:0;background:#0b0f19;padding:24px'>"
        "<table role='presentation' width='100%' cellpadding='0' cellspacing='0' "
        "style='max-width:640px;margin:0 auto;background:#ffffff;border-radius:12px;"
        "overflow:hidden;border:1px solid #e5e7eb'>"
        "<tr><td style='padding:22px 26px;background:#0b0f19;color:#ffffff;"
        "font-family:Arial,sans-serif'>"
        "<div style='font-size:11px;letter-spacing:.25em;color:#9ca3af'>"
        "JWOOD TECHNOLOGIES</div>"
        "<div style='font-size:20px;margin-top:4px'>New project inquiry</div>"
        "</td></tr>"
        f"<tr><td style='padding:0'><table width='100%' cellpadding='0' "
        f"cellspacing='0' style='border-collapse:collapse'>{rows}</table></td></tr>"
        "<tr><td style='padding:16px 26px;color:#6b7280;font-family:Arial,sans-serif;"
        "font-size:12px;border-top:1px solid #e5e7eb'>"
        f"Submitted {esc(r.created_at)}"
        "</td></tr>"
        "</table></body></html>"
    )


async def _send_inquiry_email(r: ContactRecord) -> bool:
    if not (RESEND_API_KEY and RECIPIENT_EMAIL):
        return False
    subject = f"New inquiry — {r.name} ({r.project_type})"
    html = _inquiry_html(r)
    # Try verified brand sender first, fall back to Resend sandbox if the
    # domain is not yet verified. Either way, replies route to info@.
    attempts: list[str] = []
    if SENDER_EMAIL:
        attempts.append(f"Jwood Technologies <{SENDER_EMAIL}>")
    attempts.append("Jwood Technologies <onboarding@resend.dev>")

    for sender in attempts:
        params = {
            "from": sender,
            "to": [RECIPIENT_EMAIL],
            "reply_to": r.email or SENDER_EMAIL,
            "subject": subject,
            "html": html,
        }
        try:
            await asyncio.to_thread(resend.Emails.send, params)
            logger.info("inquiry email sent via %s", sender)
            return True
        except Exception as exc:
            logger.warning("Resend send failed via %s: %s", sender, exc)
            continue
    return False


@api.post("/contact", response_model=ContactRecord)
async def create_contact(payload: ContactCreate):
    record = ContactRecord(
        name=payload.name.strip(),
        email=payload.email,
        phone=(payload.phone or "").strip(),
        project_type=payload.project_type.strip(),
        description=payload.description.strip(),
        budget=payload.budget.strip(),
        timeline=payload.timeline.strip(),
    )
    # Fire email best-effort
    sent = await _send_inquiry_email(record)
    record.email_sent = sent
    await db.contact_submissions.insert_one(record.model_dump())
    await _metric_inc("contact_submissions_count", 1)
    logger.info(
        "Contact stored %s <%s> — email_sent=%s", record.name, record.email, sent
    )
    return record


@api.get("/contact", response_model=List[ContactRecord])
async def list_contacts():
    docs = (
        await db.contact_submissions.find({}, {"_id": 0})
        .sort("created_at", -1)
        .to_list(500)
    )
    return docs


# ---------------------------------------------------------------------------
# Chatbot (structured intake on homepage)
# ---------------------------------------------------------------------------
class ChatbotCreate(BaseModel):
    first_name: str
    last_name: str
    email: EmailStr
    phone: Optional[str] = ""
    question: str


class ChatbotRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    first_name: str
    last_name: str
    email: str
    phone: str = ""
    question: str
    email_sent: bool = False
    created_at: str = Field(
        default_factory=lambda: datetime.now(timezone.utc).isoformat()
    )


def _chatbot_html(r: ChatbotRecord) -> str:
    def esc(v: str) -> str:
        return (
            (v or "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    row = (
        "<tr><td style='padding:10px 14px;background:#f6f7f9;color:#6b7280;"
        "font-family:Arial,sans-serif;font-size:12px;letter-spacing:.08em;"
        "text-transform:uppercase;width:140px;vertical-align:top'>{k}</td>"
        "<td style='padding:10px 14px;color:#0b0f19;font-family:Arial,sans-serif;"
        "font-size:14px;vertical-align:top'>{v}</td></tr>"
    )
    rows = "".join(
        row.format(k=k, v=esc(v).replace("\n", "<br>"))
        for k, v in [
            ("First name", r.first_name),
            ("Last name", r.last_name),
            ("Email", r.email),
            ("Phone", r.phone or "—"),
            ("Question", r.question),
        ]
    )
    return (
        "<!doctype html><html><body style='margin:0;background:#0b0f19;padding:24px'>"
        "<table role='presentation' width='100%' cellpadding='0' cellspacing='0' "
        "style='max-width:640px;margin:0 auto;background:#ffffff;border-radius:12px;"
        "overflow:hidden;border:1px solid #e5e7eb'>"
        "<tr><td style='padding:22px 26px;background:#0b0f19;color:#ffffff;"
        "font-family:Arial,sans-serif'>"
        "<div style='font-size:11px;letter-spacing:.25em;color:#9ca3af'>"
        "JWOOD TECHNOLOGIES · WOOD AI</div>"
        "<div style='font-size:20px;margin-top:4px'>New chatbot lead</div>"
        "</td></tr>"
        f"<tr><td style='padding:0'><table width='100%' cellpadding='0' "
        f"cellspacing='0' style='border-collapse:collapse'>{rows}</table></td></tr>"
        "<tr><td style='padding:16px 26px;color:#6b7280;font-family:Arial,sans-serif;"
        "font-size:12px;border-top:1px solid #e5e7eb'>"
        f"Submitted {esc(r.created_at)}"
        "</td></tr>"
        "</table></body></html>"
    )


async def _send_chatbot_email(r: ChatbotRecord) -> bool:
    if not (RESEND_API_KEY and RECIPIENT_EMAIL):
        return False
    subject = f"Wood AI Chatbot — {r.first_name} {r.last_name}"
    html = _chatbot_html(r)
    attempts: list[str] = []
    if SENDER_EMAIL:
        attempts.append(f"Jwood Technologies <{SENDER_EMAIL}>")
    attempts.append("Jwood Technologies <onboarding@resend.dev>")
    for sender in attempts:
        params = {
            "from": sender,
            "to": [RECIPIENT_EMAIL],
            "reply_to": r.email or SENDER_EMAIL,
            "subject": subject,
            "html": html,
        }
        try:
            await asyncio.to_thread(resend.Emails.send, params)
            logger.info("chatbot email sent via %s", sender)
            return True
        except Exception as exc:
            logger.warning("chatbot Resend send failed via %s: %s", sender, exc)
            continue
    return False


# ---------------------------------------------------------------------------
# EON — public homepage AI assistant (brand-aware + general)
# ---------------------------------------------------------------------------
class EonMessage(BaseModel):
    role: str  # "user" | "assistant"
    text: str


class EonChatIn(BaseModel):
    message: str
    history: List[EonMessage] = Field(default_factory=list)


EON_SYSTEM_PROMPT = (
    "You are EON, the AI assistant for Jwood Technologies — a premium technology firm "
    "that builds AI-native products, bespoke software, web and mobile applications, "
    "data platforms, and automation. You answer with calm confidence in 1–4 short "
    "sentences of plain prose. No markdown headings, no bullet lists unless explicitly "
    "asked. You can answer general questions about anything (science, code, writing, "
    "ideas), but when a visitor sounds like a prospective client (mentions a project, "
    "budget, timeline, hiring, services), gently steer them toward the inquiry form on "
    "this page or the contact email info@jwoodtechnologies.com. Never invent prices. "
    "Never claim to be a human. Stay sharp, honest, and useful."
)


@api.post("/eon/chat")
async def eon_chat(body: EonChatIn):
    """Public homepage EON. No auth. Brand-aware + general assistant."""
    msg = (body.message or "").strip()
    if not msg:
        raise HTTPException(status_code=400, detail="Message required.")
    if not openai_client:
        return {
            "reply": (
                "EON is in BETA and isn't connected to a model yet. Drop your "
                "details in the inquiry form below and the team will reach out."
            )
        }
    convo = [{"role": "system", "content": EON_SYSTEM_PROMPT}]
    # last 10 turns max, keeps the prompt tight
    for m in (body.history or [])[-10:]:
        role = "assistant" if m.role == "assistant" else "user"
        text = (m.text or "").strip()
        if text:
            convo.append({"role": role, "content": text[:2000]})
    convo.append({"role": "user", "content": msg[:2000]})
    try:
        resp = await openai_client.chat.completions.create(
            model=SUMMARY_MODEL,
            messages=convo,
            temperature=0.5,
            max_tokens=320,
        )
        reply = (resp.choices[0].message.content or "").strip()
        await _metric_inc("eon_chat_calls", 1)
        if resp.usage:
            await _metric_inc("openai_chat_tokens", int(resp.usage.total_tokens or 0))
            await _metric_inc("openai_chat_calls", 1)
    except Exception as exc:  # pragma: no cover - network/runtime
        logger.warning("EON chat failed: %s", exc)
        reply = "EON is unavailable right now. Please try again in a moment."
    return {"reply": reply}


@api.post("/chatbot", response_model=ChatbotRecord)
async def chatbot_intake(payload: ChatbotCreate):
    record = ChatbotRecord(
        first_name=payload.first_name.strip(),
        last_name=payload.last_name.strip(),
        email=payload.email,
        phone=(payload.phone or "").strip(),
        question=payload.question.strip(),
    )
    sent = await _send_chatbot_email(record)
    record.email_sent = sent
    await db.chatbot_submissions.insert_one(record.model_dump())
    await _metric_inc("chatbot_submissions_count", 1)
    logger.info(
        "chatbot stored %s %s <%s> email_sent=%s",
        record.first_name,
        record.last_name,
        record.email,
        sent,
    )
    return record


# ---------------------------------------------------------------------------
# Vineyard auth
# ---------------------------------------------------------------------------
@api.post("/vineyard/auth")
async def vineyard_auth(req: AuthRequest):
    if req.password.strip() != VINEYARD_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid password")
    return {"ok": True, "token": "vineyard-session"}


# ---------------------------------------------------------------------------
# Sources
# ---------------------------------------------------------------------------
@api.get("/vineyard/sources", response_model=List[SourceRecord])
async def list_sources():
    docs = (
        await db.sources.find({}, {"_id": 0}).sort("created_at", 1).to_list(200)
    )
    return docs


@api.get("/vineyard/sources/status")
async def sources_status():
    """Source list enriched with live indexed_count per source.

    Powers the Sources Status drawer AND the dynamic Source-filter chip
    row in the search UI. Reads counts only from the LOCKED active
    archive — in-flight `build-*` chunks are invisible so the user never
    sees half-crawled numbers.

    Sources flagged in `HIDDEN_SOURCE_LABELS` (e.g. retired RDA xlsx) are
    omitted from this response so they don't appear in either UI.
    """
    meta = await _get_active_index()
    active_v = meta.get("version") or "v1"
    pipeline = [
        {"$match": {"index_version": active_v}},
        {"$group": {"_id": "$source_id", "n": {"$sum": 1}}},
    ]
    counts: dict[str, int] = {}
    async for r in db.documents.aggregate(pipeline):
        if r["_id"]:
            counts[r["_id"]] = r["n"]

    raw = (
        await db.sources.find({}, {"_id": 0}).sort("created_at", 1).to_list(200)
    )
    out = []
    for s in raw:
        label = (s.get("label") or "").strip()
        if label in HIDDEN_SOURCE_LABELS:
            continue
        # Source must have at least one indexed doc to appear (otherwise
        # users see empty chips for half-broken sources).
        n = int(counts.get(s.get("id"), 0))
        if n == 0 and s.get("status") != "crawling":
            continue
        out.append(
            {
                "id": s.get("id"),
                "url": s.get("url"),
                "label": label,
                "display_name": _clean_source_name(label, s.get("url") or ""),
                "status": s.get("status") or "idle",
                "indexed_count": n,
                "created_at": s.get("created_at"),
            }
        )
    return {
        "active_version": active_v,
        "total_indexed": int(meta.get("doc_count") or 0),
        "sources": out,
    }


# Sources whose indexed content is kept in the archive but should NOT
# surface in the UI (chips, sources panel) — typically retired feeds.
HIDDEN_SOURCE_LABELS = {
    "RDA Past Meetings Index",
}


def _clean_source_name(label: str, url: str) -> str:
    """Convert raw labels like 'Vineyard Utah · Official Site' into the
    short, friendly names the user requested for filter chips."""
    label = (label or "").strip()
    low = label.lower()
    if "civicclerk" in low or "meetings" in low:
        return "CivicClerk Meetings"
    if "municipal code" in low or "municode" in low:
        return "Municipal Code"
    if "transparent" in low:
        return "Utah Transparency"
    if "utah.gov" in low or low.startswith("utah ") or "state" in low:
        return "Utah.gov"
    if "vineyard utah" in low or "vineyardutah" in low or "official" in low:
        return "Vineyard Official Site"
    # Fall back to the user-provided label, or hostname if no label.
    if label:
        # Strip trailing parenthetical "(state-wide, ...)" noise
        if "(" in label:
            label = label.split("(")[0].strip(" ·-")
        return label or "Source"
    try:
        from urllib.parse import urlparse
        return urlparse(url).netloc.replace("www.", "") or "Source"
    except Exception:
        return "Source"


@api.post("/vineyard/sources", response_model=SourceRecord)
async def add_source(payload: SourceCreate, background: BackgroundTasks):
    url = payload.url.strip()
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https") or not parsed.netloc:
        raise HTTPException(status_code=400, detail="Please enter a valid URL.")
    # normalise (strip trailing slash for dedupe)
    canonical = url.rstrip("/")
    existing = await db.sources.find_one(
        {"$or": [{"url": url}, {"url": canonical}, {"url": canonical + "/"}]},
        {"_id": 0},
    )
    if existing:
        raise HTTPException(status_code=409, detail="Source already indexed.")
    label = (payload.label or "").strip() or parsed.netloc.replace("www.", "")
    record = SourceRecord(url=url, label=label, status="crawling")
    await db.sources.insert_one(record.model_dump())
    background.add_task(crawl_source_task, record.id, url, label)
    return record


@api.delete("/vineyard/sources/{source_id}")
async def delete_source(source_id: str):
    await db.sources.delete_one({"id": source_id})
    await db.documents.delete_many({"source_id": source_id})
    return {"ok": True}


@api.post("/vineyard/sources/{source_id}/crawl")
async def trigger_crawl(source_id: str, background: BackgroundTasks):
    src = await db.sources.find_one({"id": source_id}, {"_id": 0})
    if not src:
        raise HTTPException(status_code=404, detail="Source not found")
    await db.sources.update_one(
        {"id": source_id}, {"$set": {"status": "crawling", "last_error": ""}}
    )
    background.add_task(crawl_source_task, source_id, src["url"], src.get("label", ""))
    return {"ok": True, "status": "crawling"}


# ---------------------------------------------------------------------------
# Index status + refresh
# ---------------------------------------------------------------------------
@api.get("/vineyard/search-ready")
async def vineyard_search_ready():
    """Pure read of the LOCKED archive's frozen stats. The number returned
    here NEVER reflects an in-flight crawl — it's whatever the admin last
    locked. Single source of truth is `db.index_meta`."""
    meta = await _get_active_index()
    total_docs = int(meta.get("doc_count") or 0)
    return {
        "ready": total_docs > 0,
        "total_docs": total_docs,
        "version": meta.get("version"),
        "locked_at": meta.get("locked_at"),
    }


@api.post("/vineyard/admin/rebuild-index")
async def vineyard_admin_rebuild_index(
    background: BackgroundTasks, password: str
):
    """Admin-only entry point for re-crawling the configured sources.

    Builds into a NEW `index_version` so the user-facing archive stays
    untouched and stable while the crawl runs. On full success across all
    sources the active version is atomically flipped via
    `POST /vineyard/admin/lock-index` (or by the orchestrator finishing).
    """
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid password")
    sources = await db.sources.find({}, {"_id": 0}).to_list(50)
    if not sources:
        raise HTTPException(status_code=400, detail="No sources configured")
    for s in sources:
        await db.sources.update_one(
            {"id": s["id"]},
            {"$set": {"status": "crawling", "last_error": ""}},
        )
        background.add_task(
            crawl_source_task, s["id"], s["url"], s.get("label", "")
        )
    return {"ok": True, "started": len(sources)}


@api.post("/vineyard/admin/lock-index")
async def vineyard_admin_lock_index(password: str):
    """Atomically promote the freshest `build-*` version to the active archive.
    Refuses if any source is still crawling.

    The previous active version is **retained** (renamed to `prev-<old>`) so
    the admin can roll back with one click. Older versions are cleaned up.
    """
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid password")
    in_flight = await db.sources.find(
        {"status": "crawling"}, {"_id": 0, "id": 1}
    ).to_list(50)
    if in_flight:
        raise HTTPException(
            status_code=409,
            detail="Cannot lock while a crawl is still running.",
        )
    # Pick the most recently-touched build version
    pipeline = [
        {"$match": {"index_version": {"$regex": "^build-"}}},
        {"$group": {"_id": "$index_version", "n": {"$sum": 1}}},
        {"$sort": {"n": -1}},
        {"$limit": 1},
    ]
    rows = await db.documents.aggregate(pipeline).to_list(1)
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No staged build chunks to lock. Run a crawl first.",
        )
    new_version = rows[0]["_id"]
    new_count = int(rows[0]["n"])

    # Identify the previous active version so we can RETAIN it as `prev-...`
    cur = await _get_active_index()
    prev_version = cur.get("version")
    prev_renamed = None
    if prev_version and prev_version != new_version:
        prev_renamed = f"prev-{prev_version}"
        # Purge any older retained `prev-*` so we keep at most ONE rollback
        await db.documents.delete_many(
            {"index_version": {"$regex": "^prev-"}}
        )
        await db.documents.update_many(
            {"index_version": prev_version},
            {"$set": {"index_version": prev_renamed}},
        )

    # Lock the new version
    await lock_index_version(new_version, new_count, locked_by="admin-lock")

    # Drop any stray build-*  that aren't the chosen one
    await db.documents.delete_many(
        {
            "index_version": {"$regex": "^build-"},
            "$expr": {"$ne": ["$index_version", new_version]},
        }
    )
    return {
        "ok": True,
        "version": new_version,
        "doc_count": new_count,
        "previous_version": prev_renamed,
    }


@api.post("/vineyard/admin/rollback-index")
async def vineyard_admin_rollback_index(password: str):
    """Flip the active archive back to the retained `prev-*` version.
    The current active version is dropped (since we already had the prev
    copy — there's no second-step rollback)."""
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid password")
    pipeline = [
        {"$match": {"index_version": {"$regex": "^prev-"}}},
        {"$group": {"_id": "$index_version", "n": {"$sum": 1}}},
        {"$sort": {"n": -1}},
        {"$limit": 1},
    ]
    rows = await db.documents.aggregate(pipeline).to_list(1)
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No previous version retained — rollback unavailable.",
        )
    rollback_tag = rows[0]["_id"]
    rollback_count = int(rows[0]["n"])
    # Re-tag prev → restored version name (strip the prefix)
    restored = rollback_tag[5:] or rollback_tag  # drop "prev-"
    await db.documents.update_many(
        {"index_version": rollback_tag},
        {"$set": {"index_version": restored}},
    )
    # Drop the version that WAS active (everything else that isn't the
    # restored version becomes garbage at this point).
    await db.documents.delete_many(
        {"index_version": {"$nin": [restored]}}
    )
    await lock_index_version(restored, rollback_count, locked_by="admin-rollback")
    return {
        "ok": True,
        "restored_version": restored,
        "doc_count": rollback_count,
    }


@api.get("/vineyard/admin/index-versions")
async def vineyard_admin_index_versions(password: str):
    """Inventory of every index_version present plus what's locked."""
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid password")
    rows = await db.documents.aggregate(
        [
            {
                "$group": {
                    "_id": "$index_version",
                    "n": {"$sum": 1},
                }
            },
            {"$sort": {"n": -1}},
        ]
    ).to_list(50)
    meta = await _get_active_index()
    return {
        "active_version": meta.get("version"),
        "active_locked_at": meta.get("locked_at"),
        "active_doc_count": int(meta.get("doc_count") or 0),
        "versions": [
            {
                "version": r["_id"] or "(untagged)",
                "doc_count": int(r["n"]),
                "kind": (
                    "active"
                    if r["_id"] == meta.get("version")
                    else (
                        "build"
                        if (r["_id"] or "").startswith("build-")
                        else "previous"
                        if (r["_id"] or "").startswith("prev-")
                        else "other"
                    )
                ),
            }
            for r in rows
        ],
    }


@api.get("/vineyard/index-status")
async def index_status():
    """Pure, non-destructive read. Returns the live document count and
    source rows. Never deletes data. Never triggers a crawl. The stable
    `db.documents.count_documents({})` is the single source of truth."""
    total_docs = await db.documents.count_documents({})
    sources = (
        await db.sources.find({}, {"_id": 0}).sort("created_at", 1).to_list(50)
    )
    is_crawling = any(s.get("status") == "crawling" for s in sources)
    return {
        "total_docs": total_docs,
        "sources": sources,
        "is_indexed": total_docs > 0,
        "is_crawling": is_crawling,
    }


@api.post("/vineyard/refresh-index")
async def refresh_index(background: BackgroundTasks):
    sources = await db.sources.find({}, {"_id": 0}).to_list(50)
    if not sources:
        raise HTTPException(status_code=400, detail="No sources configured")
    for s in sources:
        await db.sources.update_one(
            {"id": s["id"]},
            {"$set": {"status": "crawling", "last_error": ""}},
        )
        background.add_task(
            crawl_source_task, s["id"], s["url"], s.get("label", "")
        )
    return {"ok": True, "started": len(sources)}


@api.get("/vineyard/documents")
async def list_documents():
    docs = (
        await db.documents.find({}, {"_id": 0, "content": 0, "embedding": 0})
        .sort("created_at", -1)
        .to_list(500)
    )
    return docs


# ---------------------------------------------------------------------------
# Crawler
# ---------------------------------------------------------------------------
MCO_TYPES = [
    ("ordinances", "Ordinances"),
    ("districts", "Districts"),
    ("resolutions", "Resolutions"),
    ("plan", "General Plan"),
    ("subdivords", "Subdivision Ordinances"),
    ("orddoc", "Ordinance Documents"),
    ("minutes", "Minutes"),
    ("zoning", "Zoning"),
    ("landscaping", "Landscaping"),
]

SECTION_PATTERNS = [
    re.compile(r"\b(?:Ordinance|Ord\.)\s*(?:No\.?\s*)?[\w\-\.]+", re.IGNORECASE),
    re.compile(r"\b(?:Resolution|Res\.)\s*(?:No\.?\s*)?[\w\-\.]+", re.IGNORECASE),
    re.compile(r"\bSection\s+\d+(?:[\-\.]\d+)*", re.IGNORECASE),
    re.compile(r"\bChapter\s+\d+(?:[\-\.]\d+)*", re.IGNORECASE),
    re.compile(r"\bTitle\s+\d+(?:[\-\.]\d+)*", re.IGNORECASE),
]

USER_AGENT = (
    "Mozilla/5.0 (compatible; JwoodVineyardBot/1.0; "
    "+https://jwoodtechnologies.com/)"
)
MAX_PAGES_PER_SOURCE = int(os.environ.get("MAX_PAGES_PER_SOURCE", "5000"))
MAX_DEPTH = int(os.environ.get("MAX_CRAWL_DEPTH", "10"))
REQUEST_TIMEOUT = 30.0


def find_section_ref(text: str) -> Optional[str]:
    for pat in SECTION_PATTERNS:
        m = pat.search(text)
        if m:
            return m.group(0).strip()
    return None


def chunk_text(text: str, size: int = 1400, overlap: int = 200) -> List[str]:
    text = re.sub(r"\s+\n", "\n", text).strip()
    if not text:
        return []
    if len(text) <= size:
        return [text]
    chunks: List[str] = []
    start = 0
    while start < len(text):
        end = min(len(text), start + size)
        chunk = text[start:end]
        if end < len(text):
            last_break = max(chunk.rfind("\n\n"), chunk.rfind(". "))
            if last_break > size * 0.5:
                end = start + last_break + 1
                chunk = text[start:end]
        chunks.append(chunk.strip())
        start = end - overlap
        if start < 0:
            start = 0
        if end >= len(text):
            break
    return [c for c in chunks if c]


def html_to_text(html: str) -> tuple[str, str]:
    soup = BeautifulSoup(html, "lxml")
    title_tag = soup.find("title")
    title = title_tag.get_text(strip=True) if title_tag else ""
    if not title:
        h1 = soup.find("h1")
        if h1:
            title = h1.get_text(strip=True)
    for tag in soup(
        ["script", "style", "noscript", "nav", "header", "footer", "aside"]
    ):
        tag.decompose()
    text = soup.get_text("\n", strip=True)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return title, text


def pdf_to_text(data: bytes) -> str:
    try:
        with fitz.open(stream=data, filetype="pdf") as doc:
            parts = [page.get_text("text") for page in doc]
        return "\n\n".join(parts).strip()
    except Exception as exc:
        logger.warning("PDF parse failed: %s", exc)
        return ""


def xlsx_to_text(data: bytes) -> tuple[str, str, list[str]]:
    """Return (title, full text, ordered list of unique hyperlink URLs).

    Cells with hyperlinks are inlined as `<value> [<url>]` so BM25 retrieval
    can surface them and the URLs are captured separately for follow-up
    crawling.
    """
    try:
        # data_only=True resolves formulas; read_only=False is required to
        # access cell.hyperlink.target.
        wb = load_workbook(io.BytesIO(data), data_only=True)
        parts: list[str] = []
        title = wb.sheetnames[0] if wb.sheetnames else "Spreadsheet"
        seen_urls: set[str] = set()
        ordered_urls: list[str] = []
        for sheet in wb.worksheets:
            parts.append(f"## Sheet: {sheet.title}")
            for row in sheet.iter_rows():
                cells: list[str] = []
                for cell in row:
                    val = cell.value
                    href = (
                        cell.hyperlink.target
                        if cell.hyperlink and cell.hyperlink.target
                        else None
                    )
                    if val is None and not href:
                        continue
                    text = str(val).strip() if val is not None else ""
                    if href:
                        if href not in seen_urls:
                            seen_urls.add(href)
                            ordered_urls.append(href)
                        if text and text != href:
                            cells.append(f"{text} [{href}]")
                        else:
                            cells.append(href)
                    elif text:
                        cells.append(text)
                if cells:
                    parts.append(" | ".join(cells))
        return title, "\n".join(parts).strip(), ordered_urls
    except Exception as exc:
        logger.warning("XLSX parse failed: %s", exc)
        return "", "", []


def same_host(a: str, b: str) -> bool:
    try:
        return (
            urlparse(a).netloc.replace("www.", "")
            == urlparse(b).netloc.replace("www.", "")
        )
    except Exception:
        return False


def normalise(url: str) -> str:
    url, _ = urldefrag(url)
    return url.rstrip("/")


async def embed_texts(texts: list[str]) -> list[list[float]]:
    """Batch-embed. Returns [] for any failure so indexing continues gracefully."""
    if not openai_client or not texts:
        return [[] for _ in texts]
    out: list[list[float]] = []
    try:
        for i in range(0, len(texts), 96):
            batch = texts[i : i + 96]
            resp = await openai_client.embeddings.create(
                model=EMBED_MODEL, input=batch
            )
            out.extend([d.embedding for d in resp.data])
            await _metric_inc("openai_embed_calls", 1)
            await _metric_inc("openai_embed_tokens", resp.usage.total_tokens)
        return out
    except Exception as exc:
        logger.warning("embedding batch failed: %s", exc)
        return [[] for _ in texts]


def _derive_source_site(source_root: str, url: str) -> str:
    """Bucket a URL into one of the canonical filter values."""
    h = (url or source_root or "").lower()
    if "civicclerk" in h:
        return "civicclerk"
    if "municipalcodeonline" in h or "municode" in h:
        return "municode"
    if "vineyardutah.gov" in h or "vineyardutah.org" in h:
        return "vineyardutah"
    if "transparent.utah.gov" in h:
        return "transparent"
    if "utah.gov" in h:
        return "utahgov"
    if "rda" in h or "revize.com" in h:
        return "rda"
    return "other"


# Match yyyy-mm-dd, mm-dd-yyyy, "January 5, 2024", etc. inside titles or URLs.
_DATE_PATS = [
    (re.compile(r"(\d{4})-(\d{1,2})-(\d{1,2})"), "ymd"),
    (re.compile(r"(\d{1,2})[/_-](\d{1,2})[/_-](\d{4})"), "mdy"),
    (
        re.compile(
            r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|"
            r"jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|"
            r"nov(?:ember)?|dec(?:ember)?)\s+(\d{1,2})(?:st|nd|rd|th)?[,]?\s+(\d{4})",
            re.IGNORECASE,
        ),
        "mname",
    ),
]
_MONTH_NAMES = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _derive_meeting_date(title: str, url: str) -> Optional[str]:
    """Pull the first plausible date from the title or URL. Returns ISO."""
    for haystack in (title or "", url or ""):
        for pat, kind in _DATE_PATS:
            m = pat.search(haystack)
            if not m:
                continue
            try:
                if kind == "ymd":
                    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                elif kind == "mdy":
                    mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                else:  # mname
                    mo = _MONTH_NAMES[m.group(1)[:3].lower()]
                    d = int(m.group(2))
                    y = int(m.group(3))
                if 1990 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
                    return f"{y:04d}-{mo:02d}-{d:02d}"
            except Exception:
                continue
    return None


def _derive_doc_type(
    *, source_site: str, url: str, title: str, is_pdf: bool, is_xlsx: bool
) -> str:
    """Categorize chunks for the user-facing filter dropdown.

    Order matters — title keywords are checked FIRST so that, e.g., a
    municode "Resolution 2023-04" is tagged `resolution` instead of being
    swallowed into the generic `ordinance` bucket. This is what powers
    both the doc-type filter chips AND the search-time ranking boost
    (official documents outrank generic pages).
    """
    u = (url or "").lower()
    t = (title or "").lower()
    blob = f"{t} {u}"

    if is_xlsx:
        return "xlsx"

    # Authoritative document types — detected via title/URL keywords.
    # Order is intentional: more specific first (resolution > ordinance >
    # minutes > agenda > attachment).
    if "resolution" in blob:
        return "resolution"
    if "ordinance" in blob or " ord " in f" {blob} ":
        return "ordinance"
    if "minutes" in blob:
        return "minutes"
    if "agenda packet" in blob or "agenda" in blob:
        return "agenda"

    # Per-source defaults
    if source_site == "rda":
        return "rda"
    if source_site == "transparent":
        return "transparency"
    if source_site == "municode":
        return "ordinance"  # default municode bucket
    if source_site == "civicclerk":
        if is_pdf:
            return "attachment"
        return "agenda"

    if is_pdf:
        return "pdf"
    return "page"


# Ranking boost applied at search time. Higher = surfaces above generic
# `page` results. Tuned so that an official document always wins over a
# random web page that merely mentions the query terms.
DOC_TYPE_BOOST = {
    "resolution": 1.30,
    "ordinance": 1.25,
    "minutes": 1.20,
    "agenda": 1.18,
    "attachment": 1.10,
    "transparency": 1.10,
    "rda": 1.10,
    "pdf": 1.05,
    "xlsx": 1.05,
    "page": 1.00,
}


async def _store_chunks(
    chunks: list[str],
    *,
    source_id: str,
    source_root: str,
    source_label: str,
    url: str,
    title: str,
    pdf_url: Optional[str],
    is_pdf: bool,
    is_xlsx: bool,
    depth: int,
    crawl_run_id: Optional[str] = None,
) -> int:
    if not chunks:
        return 0
    # Atlas free tier (512 MB) cannot fit per-chunk embeddings (~12 KB each).
    # Search code already falls back to BM25-only when the field is missing,
    # so we skip the embed call entirely — saves OpenAI cost AND storage.
    site_bucket = _derive_source_site(source_root, url)
    meeting_dt = _derive_meeting_date(title, url)
    dtype = _derive_doc_type(
        source_site=site_bucket,
        url=url,
        title=title,
        is_pdf=is_pdf,
        is_xlsx=is_xlsx,
    )
    docs = []
    for chunk in chunks:
        rec = DocumentRecord(
            source_id=source_id,
            source_root=source_root,
            source_label=source_label,
            source_site=site_bucket,
            url=url,
            pdf_url=pdf_url,
            title=title,
            section_ref=find_section_ref(chunk[:2000]),
            excerpt=chunk[:220].strip(),
            content=chunk,
            is_pdf=is_pdf,
            is_xlsx=is_xlsx,
            doc_type=dtype,
            meeting_date=meeting_dt,
            depth=depth,
        )
        d = rec.model_dump()
        if crawl_run_id:
            d["crawl_run_id"] = crawl_run_id
        # Tag with a building version so reads scoped to the active archive
        # never see in-flight chunks. Active version flips only on admin swap.
        d["index_version"] = f"build-{crawl_run_id}" if crawl_run_id else "v1"
        docs.append(d)
    await db.documents.insert_many(docs)
    return len(docs)


async def _crawl_filtered_state_site(
    source_id: str,
    root_url: str,
    source_label: str = "",
    crawl_run_id: Optional[str] = None,
    *,
    mention_keyword: str = "vineyard",
    host_suffix: str = "utah.gov",
    max_fetches: int = 6000,
):
    """Crawl an entire state-government domain, indexing ONLY pages whose
    text body (or PDF text) contains `mention_keyword`. Crawl traverses
    every reachable URL across `*.{host_suffix}` so we can discover Vineyard
    mentions buried multiple clicks deep, but persistent storage is
    keyword-gated so the index stays focused and Atlas storage stays small.

    Hard cap on `max_fetches` to keep the crawl bounded.
    """
    keyword = mention_keyword.lower()
    seen: set[str] = set()
    # Use a deque so we can prioritise URLs that look promising (contain the
    # keyword or are inside a deparment-likely path).
    high_q: "collections.deque[tuple[str, int]]" = collections.deque()
    low_q: "collections.deque[tuple[str, int]]" = collections.deque()
    low_q.append((normalise(root_url), 0))

    discovered = 0
    indexed = 0
    fetched = 0

    async with httpx.AsyncClient(
        timeout=REQUEST_TIMEOUT,
        follow_redirects=True,
        headers={"User-Agent": USER_AGENT},
    ) as http:
        while (high_q or low_q) and fetched < max_fetches:
            url, depth = high_q.popleft() if high_q else low_q.popleft()
            if url in seen:
                continue
            seen.add(url)
            fetched += 1
            if fetched % 100 == 0:
                logger.info(
                    "utah.gov crawl progress: fetched=%d indexed=%d queue=%d",
                    fetched,
                    indexed,
                    len(high_q) + len(low_q),
                )
                # Live-update the source row so the admin can see progress
                await db.sources.update_one(
                    {"id": source_id},
                    {
                        "$set": {
                            "discovered": fetched,
                            "max_depth_reached": depth,
                        }
                    },
                )

            try:
                resp = await http.get(url)
            except Exception:
                continue
            if resp.status_code != 200:
                continue
            ctype = (resp.headers.get("content-type") or "").lower()
            low_url = url.lower()
            is_pdf = "pdf" in ctype or low_url.endswith(".pdf")

            text = ""
            title = url

            if is_pdf:
                try:
                    text = pdf_to_text(resp.content) or ""
                except Exception:
                    text = ""
                title = url.rsplit("/", 1)[-1] or "PDF"
            elif "text/html" in ctype or "xml" in ctype:
                try:
                    title, text = html_to_text(resp.text)
                except Exception:
                    text = ""
            else:
                continue

            if not text or len(text) < 80:
                # Still extract links from HTML even if we don't index
                if not is_pdf and "text/html" in ctype:
                    pass
                else:
                    continue

            text_lc = text.lower()
            mentions = keyword in text_lc

            # INDEX: only if the page mentions the keyword
            if mentions:
                chunks = chunk_text(text)
                added = await _store_chunks(
                    chunks,
                    source_id=source_id,
                    source_root=root_url,
                    source_label=source_label,
                    url=url,
                    title=(title or url)[:500],
                    pdf_url=url if is_pdf else None,
                    is_pdf=is_pdf,
                    is_xlsx=False,
                    depth=depth,
                    crawl_run_id=crawl_run_id,
                )
                if added:
                    indexed += 1
                    discovered += 1
                    logger.info(
                        "utah.gov INDEX (%d/%d): %s — %s",
                        indexed,
                        fetched,
                        title[:60],
                        url[:100],
                    )

            # QUEUE: HTML pages — extract anchors regardless of mention status
            # (we may need to traverse a department index page to reach a
            # Vineyard mention 2 clicks deep). Only follow URLs on
            # `*.utah.gov`. PDFs from other hosts are still followed if
            # explicitly linked.
            if not is_pdf and "text/html" in ctype and depth < MAX_DEPTH:
                try:
                    soup = BeautifulSoup(resp.text, "lxml")
                except Exception:
                    soup = None
                if soup is not None:
                    for a in soup.find_all("a", href=True):
                        href = a["href"].strip()
                        if href.startswith(("mailto:", "tel:", "javascript:", "#")):
                            continue
                        full = normalise(urljoin(url, href))
                        if not full.startswith("http"):
                            continue
                        # restrict crawl to *.utah.gov
                        host = urlparse(full).netloc.lower()
                        if not (host == host_suffix or host.endswith("." + host_suffix)):
                            continue
                        if full in seen:
                            continue
                        # Prioritise URLs that mention the keyword in URL or
                        # link text — those resolve faster.
                        anchor_text = (a.get_text(" ", strip=True) or "").lower()
                        promising = (
                            keyword in full.lower()
                            or keyword in anchor_text
                            or (mentions and depth < 2)
                        )
                        if promising:
                            high_q.append((full, depth + 1))
                        else:
                            low_q.append((full, depth + 1))
            await asyncio.sleep(0.05)

    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "discovered": fetched,
                "max_depth_reached": MAX_DEPTH,
            }
        },
    )
    logger.info(
        "utah.gov filtered crawl finished: fetched=%d indexed=%d (keyword=%s)",
        fetched,
        indexed,
        keyword,
    )


async def _crawl_transparent_utah(
    source_id: str,
    root_url: str,
    source_label: str = "",
    crawl_run_id: Optional[str] = None,
):
    """Pull every Vineyard City record from transparent.utah.gov's data API.

    The official Transparent Utah portal is a React SPA backed by a Cloud
    Run query handler. Vineyard City's `entity_id` is **611**; data is
    available for fiscal years 2014–2026. For each `(query_function,
    fiscal_year)` pair we POST `{name: 611, years: "<year>"}` and store
    the JSON response as a searchable document.
    """
    QUERY_API = "https://tu-query-handler-prod-uewlhjwsua-wm.a.run.app/"
    ENTITY_ID = 611
    ENTITY_NAME = "Vineyard City"
    # Each tool: (function name, label, deep-link, parameter shape)
    # Param shapes:
    #   "name_year": {"name": <id>, "years": "<year>"} — per-year query
    #   "entity_all_years": {"entity_id": <id>} — single call returns all years
    TOOLS = [
        (
            "getHighestPaidEmployees",
            "Highest Paid Employees",
            "https://old.transparent.utah.gov/emp.php",
            "name_year",
        ),
        (
            "getEntityRevenues",
            "Entity Revenues",
            "https://old.transparent.utah.gov/entity_details.php",
            "name_year",
        ),
        (
            "getEntitySurplusDeficit",
            "Entity Surplus / Deficit",
            "https://old.transparent.utah.gov/entity_details.php",
            "name_year",
        ),
        (
            "getEntityExpenses",
            "Entity Expenses (all years)",
            "https://old.transparent.utah.gov/entity_details.php",
            "entity_all_years",
        ),
    ]
    YEARS = list(range(2014, 2027))  # 2014..2026 inclusive

    indexed = 0
    fetched = 0

    async with httpx.AsyncClient(
        timeout=REQUEST_TIMEOUT,
        follow_redirects=True,
        headers={
            "User-Agent": USER_AGENT,
            "Origin": "https://old.transparent.utah.gov",
            "Referer": "https://old.transparent.utah.gov/",
        },
    ) as http:
        # First: confirm Vineyard City is currently in the entity catalog
        # (this is a free sanity check that also exercises the API).
        try:
            r = await http.post(
                QUERY_API,
                data={"function": "getAvailableEntitiesByYear", "parameter": "[]"},
            )
            if r.status_code == 200:
                ents = r.json()
                vy_years = sorted(
                    {
                        e["fiscal_year"]
                        for e in ents
                        if "vineyard" in (e.get("entity_name") or "").lower()
                    }
                )
                if vy_years:
                    logger.info(
                        "transparent-utah: Vineyard fiscal years available = %s",
                        vy_years,
                    )
        except Exception as exc:
            logger.warning("transparent-utah catalog probe failed: %s", exc)

        for fn, label, deep_link, shape in TOOLS:
            year_iter = YEARS if shape == "name_year" else [None]
            for yr in year_iter:
                fetched += 1
                if shape == "name_year":
                    params = json.dumps({"name": ENTITY_ID, "years": str(yr)})
                else:  # entity_all_years
                    params = json.dumps({"entity_id": ENTITY_ID})
                try:
                    r = await http.post(
                        QUERY_API,
                        data={"function": fn, "parameter": params},
                    )
                except Exception as exc:
                    logger.warning(
                        "transparent-utah %s/%s failed: %s", fn, yr, exc
                    )
                    continue
                if r.status_code != 200:
                    continue
                try:
                    rows = r.json()
                except Exception:
                    continue
                if not isinstance(rows, list) or not rows:
                    continue

                # For all-years queries, group rows by fiscal_year so each
                # year becomes its own searchable document.
                if shape == "entity_all_years":
                    by_year: dict[int, list[dict]] = {}
                    for row in rows:
                        if not isinstance(row, dict):
                            continue
                        fy = row.get("fiscal_year")
                        if fy is None:
                            continue
                        by_year.setdefault(int(fy), []).append(row)
                    sub_groups = list(by_year.items())
                else:
                    sub_groups = [(yr, rows)]

                for sub_year, sub_rows in sub_groups:
                    lines: list[str] = [
                        f"{ENTITY_NAME} — {label} — Fiscal Year {sub_year}",
                        f"Source: Transparent Utah ({fn})",
                        "",
                    ]
                    for row in sub_rows:
                        if not isinstance(row, dict):
                            continue
                        parts = []
                        for k, v in row.items():
                            if v is None or v == "":
                                continue
                            if (
                                isinstance(v, (int, float))
                                and any(
                                    token in k.lower()
                                    for token in (
                                        "wage",
                                        "salary",
                                        "benefit",
                                        "total",
                                        "amount",
                                        "expense",
                                        "revenue",
                                        "surplus",
                                        "deficit",
                                        "compensation",
                                        "paid",
                                        "payment",
                                        "net",
                                    )
                                )
                            ):
                                parts.append(f"{k}: ${v:,.2f}")
                            else:
                                parts.append(f"{k}: {v}")
                        if parts:
                            lines.append(" | ".join(parts))

                    body = "\n".join(lines)
                    title = f"{ENTITY_NAME} — {label} ({sub_year})"
                    chunks = chunk_text(body, size=2400, overlap=200)
                    added = await _store_chunks(
                        chunks,
                        source_id=source_id,
                        source_root=root_url,
                        source_label=source_label,
                        url=deep_link,
                        title=title,
                        pdf_url=None,
                        is_pdf=False,
                        is_xlsx=False,
                        depth=0,
                        crawl_run_id=crawl_run_id,
                    )
                    if added:
                        indexed += 1
                        logger.info(
                            "transparent-utah INDEX %s/%s — %d rows, %d chunks",
                            fn,
                            sub_year,
                            len(sub_rows),
                            added,
                        )
                # Politeness — Cloud Run will throttle if hammered.
                await asyncio.sleep(0.15)

    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "discovered": fetched,
                "max_depth_reached": 1,
            }
        },
    )
    logger.info(
        "transparent-utah crawl finished: api_calls=%d records_indexed=%d",
        fetched,
        indexed,
    )



async def crawl_source_task(source_id: str, root_url: str, source_label: str = ""):
    """Dispatcher: picks the right crawler based on URL host.

    Wraps the per-host crawler with:
      * Timeout watchdog (CRAWL_TIMEOUT_SECONDS)
      * Stage-then-swap: new chunks land with the new `crawl_run_id` while old
        ones stay searchable, swap on success, partial cleanup on failure.
      * Status state machine: crawling → done / error / timeout
    """
    host = urlparse(root_url).netloc.lower()
    run_id = uuid.uuid4().hex
    started_at = datetime.now(timezone.utc).isoformat()

    # Track current run on the source so the swap step knows what to keep.
    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "status": "crawling",
                "last_error": "",
                "current_run_id": run_id,
                "current_run_started_at": started_at,
                # NOTE: do NOT zero out pages_indexed / sections_indexed — those
                # numbers continue to reflect the LAST successful crawl until
                # this run completes.
            }
        },
    )

    async def _dispatch():
        if "municipalcodeonline.com" in host:
            await _crawl_mco(source_id, root_url, source_label, run_id)
        elif "civicclerk.com" in host:
            await _crawl_civicclerk(source_id, root_url, source_label, run_id)
        elif "transparent.utah.gov" in host:
            # Dedicated crawler: hits the Transparent Utah data API directly
            # rather than scraping the React SPA.
            await _crawl_transparent_utah(source_id, root_url, source_label, run_id)
        elif host == "utah.gov" or host.endswith(".utah.gov"):
            await _crawl_filtered_state_site(
                source_id,
                root_url,
                source_label,
                run_id,
                mention_keyword="vineyard",
                host_suffix="utah.gov",
            )
        else:
            await _crawl_generic(source_id, root_url, source_label, run_id)

    try:
        await asyncio.wait_for(_dispatch(), timeout=CRAWL_TIMEOUT_SECONDS)
    except asyncio.TimeoutError:
        logger.warning(
            "crawl timeout (%ds) for %s — keeping previous index intact",
            CRAWL_TIMEOUT_SECONDS,
            root_url,
        )
        # Drop any partial new chunks, leave previous index untouched.
        await db.documents.delete_many(
            {"source_id": source_id, "crawl_run_id": run_id}
        )
        await db.sources.update_one(
            {"id": source_id},
            {
                "$set": {
                    "status": "timeout",
                    "last_error": (
                        f"timeout after {CRAWL_TIMEOUT_SECONDS}s — "
                        "previous index preserved"
                    ),
                    "last_finished_at": datetime.now(timezone.utc).isoformat(),
                },
                "$unset": {"current_run_id": "", "current_run_started_at": ""},
            },
        )
        return
    except Exception as exc:
        logger.exception("dispatcher error: %s", exc)
        # Drop partial inserts; keep last good index searchable.
        await db.documents.delete_many(
            {"source_id": source_id, "crawl_run_id": run_id}
        )
        await db.sources.update_one(
            {"id": source_id},
            {
                "$set": {
                    "status": "error",
                    "last_error": (str(exc) or "crawler error")[:400],
                    "last_finished_at": datetime.now(timezone.utc).isoformat(),
                },
                "$unset": {"current_run_id": "", "current_run_started_at": ""},
            },
        )
        return

    # ---------- success: atomic swap ----------
    new_count = await db.documents.count_documents(
        {"source_id": source_id, "crawl_run_id": run_id}
    )
    if new_count == 0:
        # Crawler finished without inserting anything — keep old data so the
        # admin sees a clear failure rather than a wiped index.
        await db.sources.update_one(
            {"id": source_id},
            {
                "$set": {
                    "status": "error",
                    "last_error": "no documents found — previous index preserved",
                    "last_finished_at": datetime.now(timezone.utc).isoformat(),
                },
                "$unset": {"current_run_id": "", "current_run_started_at": ""},
            },
        )
        return

    # Build complete for this source. We do NOT delete other versions here
    # — that would clobber the locked active archive. The admin flips the
    # whole archive atomically by calling /vineyard/admin/lock-index once
    # every source has finished. We just mark this source done and move on.
    await db.documents.update_many(
        {"source_id": source_id, "crawl_run_id": run_id},
        {"$unset": {"crawl_run_id": ""}},
    )
    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "status": "done",
                "last_error": "",
                "last_crawled_at": datetime.now(timezone.utc).isoformat(),
                "last_finished_at": datetime.now(timezone.utc).isoformat(),
                "pages_indexed": new_count,
            },
            "$unset": {"current_run_id": "", "current_run_started_at": ""},
        },
    )


async def _crawl_generic(
    source_id: str,
    root_url: str,
    source_label: str = "",
    crawl_run_id: Optional[str] = None,
):
    """Recursive polite crawl within host. Handles HTML, PDF, XLSX."""
    seen: set[str] = set()
    queue: list[tuple[str, int]] = [(root_url, 0)]
    pages_indexed = 0

    headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    async with httpx.AsyncClient(
        timeout=REQUEST_TIMEOUT,
        follow_redirects=True,
        headers=headers,
    ) as http:
        while queue and pages_indexed < MAX_PAGES_PER_SOURCE:
            url, depth = queue.pop(0)
            url = normalise(url)
            if url in seen:
                continue
            seen.add(url)

            try:
                resp = await http.get(url)
            except Exception as exc:
                logger.warning("fetch failed %s: %s", url, exc)
                continue
            if resp.status_code >= 400:
                continue

            ctype = resp.headers.get("content-type", "").lower()
            low = url.lower()
            is_pdf = "application/pdf" in ctype or low.endswith(".pdf")
            is_xlsx = (
                "spreadsheetml" in ctype
                or "ms-excel" in ctype
                or low.endswith(".xlsx")
                or low.endswith(".xls")
            )

            if is_pdf:
                text = pdf_to_text(resp.content)
                if text:
                    title = url.rsplit("/", 1)[-1] or "PDF Document"
                    chunks = chunk_text(text)
                    added = await _store_chunks(
                        chunks,
                        source_id=source_id,
                        source_root=root_url,
                        source_label=source_label,
                        url=url,
                        title=title,
                        pdf_url=url,
                        is_pdf=True,
                        is_xlsx=False,
                        depth=depth,
                        crawl_run_id=crawl_run_id,
                    )
                    if added:
                        pages_indexed += 1
                continue

            if is_xlsx:
                title, text, xlsx_links = xlsx_to_text(resp.content)
                if text:
                    title = title or url.rsplit("/", 1)[-1] or "Spreadsheet"
                    chunks = chunk_text(text, size=1800, overlap=200)
                    added = await _store_chunks(
                        chunks,
                        source_id=source_id,
                        source_root=root_url,
                        source_label=source_label,
                        url=url,
                        title=title,
                        pdf_url=None,
                        is_pdf=False,
                        is_xlsx=True,
                        depth=depth,
                        crawl_run_id=crawl_run_id,
                    )
                    if added:
                        pages_indexed += 1
                # Follow every unique hyperlink the spreadsheet contains.
                # Skip URLs already covered by other indexed sources (e.g.
                # CivicClerk events the deep crawler grabbed) — the
                # spreadsheet is a meeting index, so most cells link there.
                if xlsx_links:
                    active_meta = await _get_active_index()
                    active_v = active_meta.get("version") or "v1"
                    for link in xlsx_links:
                        norm = normalise(link)
                        if norm in seen:
                            continue
                        # Skip CivicClerk event/* URLs — already in archive
                        if "civicclerk.com" in norm:
                            continue
                        already = await db.documents.find_one(
                            {"index_version": active_v, "url": norm},
                            {"_id": 0, "id": 1},
                        )
                        if already:
                            continue
                        # Queue the hyperlink for the same generic crawler
                        queue.append((norm, depth + 1))
                continue

            if "text/html" not in ctype and "xml" not in ctype:
                continue

            title, text = html_to_text(resp.text)
            if text and len(text) > 80:
                chunks = chunk_text(text)
                # find first PDF link on page for context
                soup_ctx = BeautifulSoup(resp.text, "lxml")
                pdf_link = None
                for a in soup_ctx.find_all("a", href=True):
                    href = a["href"]
                    if href.lower().endswith(".pdf"):
                        pdf_link = urljoin(url, href)
                        break
                added = await _store_chunks(
                    chunks,
                    source_id=source_id,
                    source_root=root_url,
                    source_label=source_label,
                    url=url,
                    title=title or url,
                    pdf_url=pdf_link,
                    is_pdf=False,
                    is_xlsx=False,
                    depth=depth,
                    crawl_run_id=crawl_run_id,
                )
                if added:
                    pages_indexed += 1

            # enqueue children (recursively, within host)
            if depth < MAX_DEPTH:
                soup = BeautifulSoup(resp.text, "lxml")
                for a in soup.find_all("a", href=True):
                    href = a["href"].strip()
                    if href.startswith(("mailto:", "tel:", "javascript:")):
                        continue
                    full = normalise(urljoin(url, href))
                    if not full.startswith("http"):
                        continue
                    # allow docs from other hosts if they are direct files
                    low2 = full.lower()
                    is_file = low2.endswith((".pdf", ".xlsx", ".xls"))
                    if not is_file and not same_host(full, root_url):
                        continue
                    if full in seen:
                        continue
                    queue.append((full, depth + 1))

            await asyncio.sleep(0.1)

    # Telemetry only — final status + pages_indexed handled by dispatcher swap.
    pdf_count = await db.documents.count_documents(
        {"source_id": source_id, "crawl_run_id": crawl_run_id, "is_pdf": True}
    ) if crawl_run_id else 0
    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "discovered": len(seen),
                "pdfs_indexed": pdf_count,
                "max_depth_reached": MAX_DEPTH,
                "failed": max(0, len(seen) - pages_indexed),
            }
        },
    )
    logger.info(
        "generic crawl finished %s pages=%d", root_url, pages_indexed
    )


# ---------------------------------------------------------------------------
# CivicClerk (JS SPA) — best-effort via Playwright
# ---------------------------------------------------------------------------
async def _crawl_civicclerk(
    source_id: str,
    root_url: str,
    source_label: str = "",
    crawl_run_id: Optional[str] = None,
):
    """Deep CivicClerk crawler using the public OData REST API.

    Strategy
    --------
    1. Page through `/v1/Events` (past + upcoming) to enumerate every meeting.
    2. For each event, hit `/v1/Events/{eventId}` → `agendaId`.
    3. Fetch `/v1/Meetings/{agendaId}` → `publishedFiles[]` + `items[]`.
    4. For every published file, request the `plainText=true` stream and
       store its text. Each file becomes its own searchable document with a
       deep-link back to the public portal page.
    5. As a fallback we also index the agenda/minutes outline (titles +
       descriptions) so meetings without published files still produce a
       searchable record.
    """
    parsed = urlparse(root_url)
    portal_host = parsed.netloc  # e.g. vineyardut.portal.civicclerk.com
    api_host = portal_host.replace(".portal.", ".api.")
    if not api_host.endswith(".civicclerk.com"):
        # belt-and-braces fallback
        api_host = "vineyardut.api.civicclerk.com"
    api_base = f"https://{api_host}/v1"

    discovered = 0
    indexed = 0

    headers = {"User-Agent": USER_AGENT, "Accept": "application/json"}
    async with httpx.AsyncClient(
        timeout=REQUEST_TIMEOUT,
        follow_redirects=True,
        headers=headers,
    ) as http:

        # --- Enumerate every event (paginated; the API caps each page at 15)
        events: list[dict] = []
        for direction in ("desc", "asc"):
            skip = 0
            empty_streak = 0
            while True:
                params = {
                    "$orderby": f"startDateTime {direction}",
                    "$top": "100",  # server still caps at ~15 per page
                    "$skip": str(skip),
                }
                try:
                    resp = await http.get(f"{api_base}/Events", params=params)
                    resp.raise_for_status()
                    page = resp.json().get("value", [])
                except Exception as exc:
                    logger.warning(
                        "civicclerk events page failed (%s skip=%d): %s",
                        direction,
                        skip,
                        exc,
                    )
                    break
                if not page:
                    break
                events.extend(page)
                advance = max(len(page), 15)
                skip += advance
                # Stop if we've passed a generous safety cap.
                if skip > 5000:
                    break
                # If we got far fewer than expected for two pages in a row,
                # assume we've drained this direction.
                if len(page) < 5:
                    empty_streak += 1
                    if empty_streak >= 2:
                        break
                else:
                    empty_streak = 0

        # Dedupe events by id (asc + desc lists may overlap on `today`)
        seen_ids: set[int] = set()
        unique_events: list[dict] = []
        for ev in events:
            eid = ev.get("id")
            if eid is None or eid in seen_ids:
                continue
            seen_ids.add(eid)
            unique_events.append(ev)
        discovered = len(unique_events)
        logger.info(
            "civicclerk: %d unique events discovered for %s",
            discovered,
            api_base,
        )

        # --- For each event, fetch its agenda + files and index everything
        for ev in unique_events:
            event_id = ev.get("id")
            agenda_id = ev.get("agendaId")
            event_name = (ev.get("eventName") or "Meeting").strip()
            event_date = (ev.get("startDateTime") or "")[:10]
            category = (ev.get("categoryName") or "").strip()
            portal_url = f"https://{portal_host}/event/{event_id}/files"

            # 1) Outline + published files
            published_files: list[dict] = []
            outline_text = ""
            if agenda_id:
                try:
                    mr = await http.get(f"{api_base}/Meetings/{agenda_id}")
                    if mr.status_code == 200:
                        meeting = mr.json()
                        published_files = meeting.get("publishedFiles") or []
                        # Build a readable outline from agenda items.
                        bits: list[str] = []
                        for it in meeting.get("items") or []:
                            num = (it.get("agendaObjectItemOutlineNumber") or "").strip()
                            name = (it.get("agendaObjectItemName") or "").strip()
                            desc = (it.get("agendaObjectItemDescription") or "").strip()
                            if name:
                                bits.append(f"{num} {name}".strip())
                            if desc:
                                bits.append(desc)
                        outline_text = "\n".join(bits).strip()
                except Exception as exc:
                    logger.warning(
                        "civicclerk meeting fetch failed event=%s agenda=%s: %s",
                        event_id,
                        agenda_id,
                        exc,
                    )

            # 2) Index the outline as the "page" record for this meeting
            meta_lines = [event_name]
            if event_date:
                meta_lines.append(f"Meeting Date: {event_date}")
            if category:
                meta_lines.append(f"Category: {category}")
            full_outline = "\n".join(meta_lines)
            if outline_text:
                full_outline = f"{full_outline}\n\nAGENDA OUTLINE\n{outline_text}"
            if len(full_outline) > 200:
                chunks = chunk_text(full_outline)
                added = await _store_chunks(
                    chunks,
                    source_id=source_id,
                    source_root=root_url,
                    source_label=source_label,
                    url=portal_url,
                    title=f"{event_name} — {event_date}".strip(" —"),
                    pdf_url=None,
                    is_pdf=False,
                    is_xlsx=False,
                    depth=1,
                    crawl_run_id=crawl_run_id,
                )
                if added:
                    indexed += 1

            # 3) Index every published file (agenda PDF, minutes, packets, …)
            for pf in published_files:
                file_id = pf.get("fileId")
                if not file_id:
                    continue
                ftype = (pf.get("type") or "File").strip()
                fname = (pf.get("name") or f"{ftype} {file_id}").strip()
                stream_pdf = (
                    f"{api_base}/Meetings/GetMeetingFileStream"
                    f"(fileId={file_id},plainText=false)"
                )
                stream_txt = (
                    f"{api_base}/Meetings/GetMeetingFileStream"
                    f"(fileId={file_id},plainText=true)"
                )
                try:
                    fr = await http.get(stream_txt)
                except Exception as exc:
                    logger.warning(
                        "civicclerk file fetch failed event=%s file=%s: %s",
                        event_id,
                        file_id,
                        exc,
                    )
                    continue
                text = ""
                if fr.status_code == 200:
                    text = (fr.text or "").strip()
                # Fallback for older meetings whose plainText extraction is
                # empty: pull the raw PDF and run PyMuPDF locally.
                if len(text) < 80:
                    try:
                        pr = await http.get(stream_pdf)
                        if (
                            pr.status_code == 200
                            and "pdf" in (pr.headers.get("content-type", "")).lower()
                        ):
                            text = (pdf_to_text(pr.content) or "").strip()
                    except Exception as exc:
                        logger.warning(
                            "civicclerk pdf fallback failed event=%s file=%s: %s",
                            event_id,
                            file_id,
                            exc,
                        )
                if len(text) < 80:
                    continue
                # Prepend a meeting header so retrieval surfaces context.
                header = f"{event_name} — {event_date} — {ftype}: {fname}"
                body = f"{header}\n\n{text}"
                chunks = chunk_text(body)
                added = await _store_chunks(
                    chunks,
                    source_id=source_id,
                    source_root=root_url,
                    source_label=source_label,
                    url=portal_url,
                    title=f"{event_name} — {ftype} — {fname}".strip(),
                    pdf_url=stream_pdf,
                    is_pdf=True,  # underlying asset is a PDF
                    is_xlsx=False,
                    depth=2,
                    crawl_run_id=crawl_run_id,
                )
                if added:
                    indexed += 1

    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "discovered": discovered,
                "max_depth_reached": 2,
            }
        },
    )
    logger.info(
        "civicclerk deep crawl finished %s events=%d indexed=%d",
        root_url,
        discovered,
        indexed,
    )


# ---------------------------------------------------------------------------
# MunicipalCodeOnline (AngularJS SPA) — deep crawler via Playwright
# ---------------------------------------------------------------------------
MCO_SECTION_SPLIT = re.compile(
    r"<div\s+class=['\"]phx-name\s*['\"][^>]*>",
    re.IGNORECASE,
)


def _mco_parse_text_blob(text_html: str, parent_nameid: str) -> list[dict]:
    """Split MCO `Text` HTML into granular sections (one per phx-name marker).

    Returns list of {heading, nameid, text}.
    """
    if not text_html:
        return []
    # Normalise <br> → newline so BeautifulSoup text is readable
    text_html = re.sub(r"<br\s*/?>", "\n", text_html, flags=re.IGNORECASE)
    # Split on phx-name markers preserving the marker
    parts = MCO_SECTION_SPLIT.split(text_html)
    out: list[dict] = []
    for i, piece in enumerate(parts):
        if not piece.strip():
            continue
        soup = BeautifulSoup(piece, "lxml")
        # Find first #name= anchor — that's the section NameId
        nameid = parent_nameid
        first_link = soup.find("a", href=re.compile(r"#name="))
        if first_link and first_link.get("href"):
            nameid = first_link["href"].split("#name=", 1)[-1].split("&")[0]
        heading_el = soup.find(["h1", "h2", "h3", "h4"]) or first_link
        heading = (heading_el.get_text(strip=True) if heading_el else "").strip()
        text = soup.get_text("\n", strip=True)
        text = re.sub(r"\n{3,}", "\n\n", text)
        if len(text) < 40:
            continue
        out.append(
            {
                "heading": heading or nameid,
                "nameid": nameid,
                "text": text,
            }
        )
    return out


async def _crawl_mco(
    source_id: str,
    root_url: str,
    source_label: str = "",
    crawl_run_id: Optional[str] = None,
):
    """Deep crawl of municipalcodeonline.com using Playwright + click-driven XHRs."""
    discovered = 0
    indexed = 0
    failed = 0
    max_depth = 0
    total_stored = 0

    parsed = urlparse(root_url)
    origin = f"{parsed.scheme}://{parsed.netloc}"

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        ctx = await browser.new_context(user_agent=USER_AGENT)
        page = await ctx.new_page()

        for type_key, type_label in MCO_TYPES:
            book_url = f"{origin}/book?type={type_key}"
            expand_body_holder: dict[str, str] = {}
            content_bodies: dict[str, str] = {}

            async def on_resp(resp, _tk=type_key):  # bind type_key
                url = resp.url
                try:
                    if "/book/expand" in url and resp.status == 200:
                        expand_body_holder["body"] = await resp.text()
                    elif "/book/content" in url and resp.status == 200:
                        from urllib.parse import urlparse as _up, parse_qs

                        q = parse_qs(_up(url).query)
                        nm = q.get("name", ["?"])[0]
                        t = q.get("type", ["?"])[0]
                        if t == _tk and nm not in content_bodies:
                            content_bodies[nm] = await resp.text()
                except Exception:
                    pass

            page.on("response", on_resp)
            try:
                await page.goto(
                    book_url, wait_until="networkidle", timeout=60000
                )
                await page.wait_for_timeout(2500)
            except Exception as exc:
                failed += 1
                logger.warning("mco goto %s failed: %s", book_url, exc)
                page.remove_listener("response", on_resp)
                continue

            # Parse top-level tree
            tree: list[dict] = []
            try:
                tree = json.loads(expand_body_holder.get("body", "[]"))
            except Exception:
                tree = []
            discovered += len(tree)

            # Click each node to trigger /book/content
            for node in tree:
                nm = node.get("NameId") or node.get("name", "")
                label = node.get("name", nm)
                if not label:
                    continue
                if nm in content_bodies:
                    continue
                try:
                    await page.click(f'text="{label}"', timeout=4000)
                    await page.wait_for_timeout(900)
                except Exception:
                    failed += 1
                    continue

            page.remove_listener("response", on_resp)

            # For each captured content body, split + store
            for nm, raw_json in content_bodies.items():
                try:
                    j = json.loads(raw_json)
                except Exception:
                    failed += 1
                    continue
                text_html = j.get("Text") or j.get("text") or ""
                if not text_html:
                    continue
                sections = _mco_parse_text_blob(text_html, nm)
                if not sections:
                    # fallback: chunk the whole blob
                    soup = BeautifulSoup(text_html, "lxml")
                    raw = soup.get_text("\n", strip=True)
                    chunks = chunk_text(raw)
                    url = f"{origin}/book?type={type_key}#name={nm}"
                    added = await _store_chunks(
                        chunks,
                        source_id=source_id,
                        source_root=root_url,
                        source_label=source_label or type_label,
                        url=url,
                        title=f"{type_label} · {nm.replace('_',' ')}",
                        pdf_url=None,
                        is_pdf=False,
                        is_xlsx=False,
                        depth=1,
                        crawl_run_id=crawl_run_id,
                    )
                    total_stored += added
                    indexed += 1 if added else 0
                    continue

                for sec in sections:
                    sec_nameid = sec["nameid"]
                    url = f"{origin}/book?type={type_key}#name={sec_nameid}"
                    title = f"{sec['heading']}" if sec["heading"] else sec_nameid
                    text = sec["text"]
                    chunks = chunk_text(text, size=1400, overlap=200)
                    added = await _store_chunks(
                        chunks,
                        source_id=source_id,
                        source_root=root_url,
                        source_label=source_label or type_label,
                        url=url,
                        title=f"{type_label} · {title}",
                        pdf_url=None,
                        is_pdf=False,
                        is_xlsx=False,
                        depth=2,
                        crawl_run_id=crawl_run_id,
                    )
                    total_stored += added
                    if added:
                        indexed += 1
                        max_depth = max(max_depth, 2)

        await browser.close()

    # Telemetry only — final status + pages_indexed handled by dispatcher swap.
    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "discovered": discovered,
                "failed": failed,
                "max_depth_reached": max_depth or 2,
            }
        },
    )
    logger.info(
        "mco crawl done %s: discovered=%d indexed_sections=%d total_stored=%d failed=%d",
        root_url,
        discovered,
        indexed,
        total_stored,
        failed,
    )


# ---------------------------------------------------------------------------
# Search — hybrid BM25 + embedding, short AI summary
# ---------------------------------------------------------------------------
WORD_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9\-]+")


def tokenize(text: str) -> list[str]:
    return [w.lower() for w in WORD_RE.findall(text)]


def cosine(a: list[float], b: list[float]) -> float:
    if not a or not b or len(a) != len(b):
        return 0.0
    dot = 0.0
    na = 0.0
    nb = 0.0
    for x, y in zip(a, b):
        dot += x * y
        na += x * x
        nb += y * y
    denom = math.sqrt(na) * math.sqrt(nb)
    if denom <= 0:
        return 0.0
    return dot / denom


async def _short_summary(query: str, top_docs: list[dict]) -> str:
    """Plain-language 2–4 sentence summary of what the top documents say.

    Strict prompt rules — the model must ground every claim in the supplied
    sources, must NOT speculate about approval/adoption status unless the
    document text explicitly states it, and must NOT add legal opinions.
    """
    if not openai_client or not top_docs:
        return ""
    ctx_blocks = []
    for i, d in enumerate(top_docs[:5], start=1):
        ctx_blocks.append(
            f"[S{i}] {d.get('title', '')}\n"
            f"Source: {d.get('source_label') or d.get('url')}\n"
            f"Type: {d.get('doc_type') or 'page'}\n"
            f"{d['content'][:1100]}"
        )
    ctx = "\n\n".join(ctx_blocks)
    system = (
        "You summarise municipal document search results. "
        "Use ONLY the supplied sources. Do not invent section numbers, "
        "ordinance numbers, dates, or facts that are not present in the "
        "text. Stay neutral — do NOT editorialise, do NOT add legal "
        "advice, do NOT speculate. "
        "If a document explicitly states it was adopted, approved, "
        "passed, or denied, you may report that. If approval status is "
        "not stated, do NOT say whether it was approved or rejected — "
        "just describe what the document contains. "
        "Write 2–4 plain sentences. No headings, no lists, no filler."
    )
    user = (
        f"Question: {query}\n\n"
        f"Top sources from the Vineyard archive:\n{ctx}\n\n"
        "Plain summary of what these sources actually say:"
    )
    try:
        resp = await openai_client.chat.completions.create(
            model=SUMMARY_MODEL,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            temperature=0.1,
            max_tokens=220,
        )
        await _metric_inc("openai_chat_calls", 1)
        if resp.usage:
            await _metric_inc("openai_chat_tokens", resp.usage.total_tokens)
        return (resp.choices[0].message.content or "").strip()
    except Exception as exc:
        logger.warning("summary failed: %s", exc)
        return ""


REFUSAL = "No clear source was found in the indexed documents."


def _build_filter(req: "SearchRequest") -> dict:
    """Translate user filter inputs into a Mongo subfilter."""
    f: dict = {}
    # Source filter — `source_id` takes precedence (set by the dynamic
    # chip row); `site` is legacy for direct-link compatibility.
    if req.source_id:
        f["source_id"] = req.source_id
    elif req.site:
        f["source_site"] = req.site
    if req.doc_type:
        f["doc_type"] = req.doc_type
    if req.date_from or req.date_to:
        rng: dict = {}
        if req.date_from:
            rng["$gte"] = req.date_from
        if req.date_to:
            rng["$lte"] = req.date_to
        f["meeting_date"] = rng
    # Hide retired sources from search results too — by id, looked up at
    # request time so any future hide additions take effect immediately.
    return f


async def _hidden_source_ids() -> list[str]:
    rows = await db.sources.find(
        {"label": {"$in": list(HIDDEN_SOURCE_LABELS)}},
        {"_id": 0, "id": 1},
    ).to_list(20)
    return [r["id"] for r in rows if r.get("id")]


def _highlight_snippet(content: str, query: str, span: int = 280) -> str:
    """Return a ~280-char window around the first query-token hit. The
    matched terms are wrapped with `<mark>...</mark>` so the frontend can
    render them bolded. Falls back to the start of the content."""
    text = re.sub(r"\s+", " ", content or "").strip()
    if not text:
        return ""
    tokens = [t for t in re.findall(r"[\w]+", query.lower()) if len(t) > 1]
    lower = text.lower()
    pos = -1
    for tok in tokens:
        i = lower.find(tok)
        if i != -1 and (pos == -1 or i < pos):
            pos = i
    if pos == -1:
        snippet = text[:span]
    else:
        start = max(0, pos - span // 3)
        snippet = text[start : start + span]
        if start > 0:
            snippet = "…" + snippet
        if start + span < len(text):
            snippet = snippet + "…"
    if tokens:
        pat = re.compile(
            r"(" + "|".join(re.escape(t) for t in tokens) + r")",
            re.IGNORECASE,
        )
        snippet = pat.sub(r"<mark>\1</mark>", snippet)
    return snippet


@api.post("/vineyard/search-all")
async def vineyard_search_all(req: SearchRequest, page: int = 1, limit: int = 20):
    """All-Documents mode: paginated raw results from the locked archive.

    Uses BM25 ranking only (no embeddings — fast). Returns one row per
    distinct URL with a ts_headline-style highlighted snippet. Filters
    apply: site, doc_type, date range.
    """
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query is required")
    try:
        return await _vineyard_search_all_impl(req, query, page, limit)
    except Exception as exc:
        logger.exception("vineyard_search_all failed: %s", exc)
        return {"total": 0, "page": page, "limit": limit, "results": []}


async def _vineyard_search_all_impl(req: SearchRequest, query: str, page: int, limit: int):
    page = max(1, int(page))
    limit = max(1, min(50, int(limit)))

    active_filter = await get_active_filter()
    user_filter = _build_filter(req)
    mongo_filter = {**active_filter, **user_filter}
    if not req.source_id:
        hidden = await _hidden_source_ids()
        if hidden:
            mongo_filter["source_id"] = {"$nin": hidden}

    # Same two-stage retrieval as /vineyard/search but we keep more
    # candidates because this endpoint is paginated.
    docs: list[dict] = []
    text_q = " ".join(t for t in tokenize(query) if len(t) >= 2)[:500]
    if text_q:
        try:
            docs = (
                await db.documents.find(
                    {**mongo_filter, "$text": {"$search": text_q}},
                    {"_id": 0, "embedding": 0, "score": {"$meta": "textScore"}},
                )
                .sort([("score", {"$meta": "textScore"})])
                .limit(2000)
                .to_list(2000)
            )
        except Exception as exc:
            logger.warning("text-search prefilter (search-all) failed: %s", exc)
            docs = []
    if not docs:
        docs = (
            await db.documents.find(
                mongo_filter, {"_id": 0, "embedding": 0}
            ).to_list(5000)
        )
    if not docs:
        return {"total": 0, "page": page, "limit": limit, "results": []}

    corpus_tokens = [tokenize(d["content"]) for d in docs]
    bm25 = BM25Okapi(corpus_tokens)
    q_tokens = tokenize(query)
    if not q_tokens:
        raise HTTPException(status_code=400, detail="Query too short")
    bm_scores = bm25.get_scores(q_tokens)

    # Score & dedupe by URL — keep best chunk per URL
    by_url: dict[str, tuple[float, dict]] = {}
    for d, score in zip(docs, bm_scores):
        if score < 0.5:
            continue
        cur = by_url.get(d["url"])
        if cur is None or score > cur[0]:
            by_url[d["url"]] = (float(score), d)
    ranked = sorted(by_url.values(), key=lambda x: x[0], reverse=True)
    total = len(ranked)
    start = (page - 1) * limit
    page_rows = ranked[start : start + limit]

    results = []
    for score, d in page_rows:
        results.append(
            {
                "id": d["id"],
                "title": d.get("title") or d["url"],
                "url": d["url"],
                "pdf_url": d.get("pdf_url"),
                "source_site": d.get("source_site", "other"),
                "source_label": d.get("source_label", ""),
                "doc_type": d.get("doc_type", "page"),
                "meeting_date": d.get("meeting_date"),
                "section_ref": d.get("section_ref"),
                "snippet": _highlight_snippet(d["content"], query),
                "score": score,
            }
        )
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "results": results,
    }


@api.get("/vineyard/stats")
async def vineyard_stats():
    """Filter-friendly stats: counts by site, doc_type, and date range."""
    active_filter = await get_active_filter()
    pipeline_site = [
        {"$match": active_filter},
        {"$group": {"_id": "$source_site", "n": {"$addToSet": "$url"}}},
    ]
    rows = await db.documents.aggregate(pipeline_site).to_list(20)
    by_site = {r["_id"] or "other": len(r["n"]) for r in rows}
    pipeline_type = [
        {"$match": active_filter},
        {"$group": {"_id": "$doc_type", "n": {"$addToSet": "$url"}}},
    ]
    rows2 = await db.documents.aggregate(pipeline_type).to_list(20)
    by_type = {r["_id"] or "page": len(r["n"]) for r in rows2}
    meta = await _get_active_index()
    return {
        "total_indexed": int(meta.get("doc_count") or 0),
        "version": meta.get("version"),
        "locked_at": meta.get("locked_at"),
        "by_site": by_site,
        "by_doc_type": by_type,
    }


@api.post("/vineyard/search", response_model=SearchResponse)
async def vineyard_search(req: SearchRequest):
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query is required")
    await _metric_inc("vineyard_search_count", 1)
    try:
        return await _vineyard_search_impl(req, query)
    except Exception as exc:
        # NEVER let the search endpoint return a 5xx — the user sees that
        # as a frightening "Search failed" toast. Log the error and return
        # an empty result with a graceful message instead.
        logger.exception("vineyard_search failed: %s", exc)
        return SearchResponse(
            answer=(
                "Search hit an internal error. Please try a slightly "
                "different query, or wait a moment and try again."
            ),
            citations=[],
            has_results=False,
            query=query,
        )


async def _vineyard_search_impl(req: SearchRequest, query: str) -> SearchResponse:
    # Read only from the locked active archive — partial in-flight crawl
    # chunks are invisible because they carry a different `index_version`.
    active_filter = await get_active_filter()
    user_filter = _build_filter(req)
    mongo_filter = {**active_filter, **user_filter}
    # Exclude retired sources unless user is explicitly filtering by one.
    if not req.source_id:
        hidden = await _hidden_source_ids()
        if hidden:
            mongo_filter["source_id"] = {"$nin": hidden}

    # Two-stage retrieval: (1) Mongo `$text` index narrows ~46k → top ~500
    # candidates by textScore, (2) BM25 + cosine + doc-type boost rerank.
    # Falls back to a streaming scan if the text query rejects (e.g. all
    # stop words) so we never lose results.
    docs: list[dict] = []
    text_q = " ".join(t for t in tokenize(query) if len(t) >= 2)[:500]
    if text_q:
        try:
            docs = (
                await db.documents.find(
                    {**mongo_filter, "$text": {"$search": text_q}},
                    {"_id": 0, "embedding": 0, "score": {"$meta": "textScore"}},
                )
                .sort([("score", {"$meta": "textScore"})])
                .limit(500)
                .to_list(500)
            )
        except Exception as exc:
            logger.warning("text-search prefilter failed: %s", exc)
            docs = []
    if not docs:
        # Fallback: tighter scan, capped to 5k
        docs = (
            await db.documents.find(
                mongo_filter, {"_id": 0, "embedding": 0}
            ).to_list(5000)
        )
    if not docs:
        return SearchResponse(
            answer=REFUSAL, citations=[], has_results=False, query=query
        )

    corpus_tokens = [tokenize(d["content"]) for d in docs]
    bm25 = BM25Okapi(corpus_tokens)
    q_tokens = tokenize(query)
    if not q_tokens:
        raise HTTPException(status_code=400, detail="Query too short")

    bm_scores = bm25.get_scores(q_tokens)
    # Top 30 candidates for rerank
    idx_scored = sorted(enumerate(bm_scores), key=lambda x: x[1], reverse=True)[:30]
    # Only refuse if NOTHING matched at all. Previously we refused at
    # bm < 1.0 which threw away long-tail matches; the new floor is 0.0
    # so any token-overlap candidate gets reranked + summarized.
    if not idx_scored or idx_scored[0][1] <= 0.0:
        return SearchResponse(
            answer=REFUSAL, citations=[], has_results=False, query=query
        )

    # Fetch embeddings for candidates
    candidate_docs = [docs[i] for i, _ in idx_scored]
    candidate_ids = [d["id"] for d in candidate_docs]
    emb_rows = await db.documents.find(
        {**mongo_filter, "id": {"$in": candidate_ids}},
        {"_id": 0, "id": 1, "embedding": 1},
    ).to_list(len(candidate_ids))
    emb_by_id = {r["id"]: r.get("embedding") or [] for r in emb_rows}

    # Query embedding (single call). If OpenAI is unavailable or the
    # request fails (e.g. no API key in production env vars), fall back to
    # BM25-only ranking. The user still gets relevant results — they just
    # won't get the AI-generated summary header.
    q_emb: list[float] = []
    try:
        q_emb_list = await embed_texts([query])
        q_emb = q_emb_list[0] if q_emb_list else []
    except Exception as exc:
        logger.warning("query embed failed (BM25 fallback): %s", exc)

    max_bm = max(s for _, s in idx_scored) or 1.0
    scored: list[tuple[float, dict, float, float]] = []
    for (i, bm), d in zip(idx_scored, candidate_docs):
        emb = emb_by_id.get(d["id"]) or []
        cos = cosine(q_emb, emb) if emb else 0.0
        # Combined: lean on embedding if available
        combined = (0.45 * (bm / max_bm)) + (0.55 * max(cos, 0.0))
        # Doc-type boost: official documents (resolution/ordinance/minutes/
        # agenda) outrank generic web pages. Multiplicative so a strong
        # text match still wins over a weak match against a boosted type.
        boost = DOC_TYPE_BOOST.get(d.get("doc_type") or "page", 1.0)
        combined *= boost
        scored.append((combined, d, bm, cos))
    scored.sort(key=lambda x: x[0], reverse=True)

    # Take top 5 distinct URLs
    top: list[tuple[float, dict, float, float]] = []
    seen_urls: set[str] = set()
    for item in scored:
        url = item[1]["url"]
        if url in seen_urls:
            continue
        seen_urls.add(url)
        top.append(item)
        if len(top) >= 5:
            break

    citations: list[Citation] = []
    for combined, d, bm, _cos in top:
        excerpt = re.sub(r"\s+", " ", d["content"])[:280].strip()
        citations.append(
            Citation(
                title=d.get("title") or d["url"],
                source_label=d.get("source_label") or urlparse(d["url"]).netloc,
                source_site=d.get("source_site", "other"),
                doc_type=d.get("doc_type", "page"),
                meeting_date=d.get("meeting_date"),
                url=d["url"],
                pdf_url=d.get("pdf_url"),
                section_ref=d.get("section_ref"),
                excerpt=excerpt,
                score=float(combined),
            )
        )

    # ALWAYS summarise top results when we have them. Only fall back to a
    # neutral default if the model itself returned an empty/refusal string
    # — never claim "no source" while citations exist below.
    answer = await _short_summary(query, [t[1] for t in top])
    if not answer or REFUSAL.lower() in answer.lower():
        # Build a deterministic 1-line description of what the top result
        # is so the user always gets a sensible header above their cites.
        top_d = top[0][1]
        dtype_label = (top_d.get("doc_type") or "page").replace("_", " ").title()
        title = top_d.get("title") or top_d.get("url")
        src = top_d.get("source_label") or "the Vineyard archive"
        answer = (
            f"Top match: {dtype_label} — “{title}” from {src}. "
            f"See the cited sources below for the full text."
        )

    return SearchResponse(
        answer=answer, citations=citations, has_results=True, query=query
    )


# ---------------------------------------------------------------------------
# Web Deep Search — broader web via DuckDuckGo + gpt-4o-mini summary
# ---------------------------------------------------------------------------
async def _ddg_search(query: str, n: int = 6) -> list[dict]:
    def _q() -> list[dict]:
        try:
            with DDGS() as d:
                return list(d.text(query, max_results=n))
        except Exception as exc:
            logger.warning("DDG search failed: %s", exc)
            return []

    return await asyncio.to_thread(_q)


@api.post("/vineyard/web-search", response_model=SearchResponse)
async def vineyard_web_search(req: SearchRequest):
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query is required")
    await _metric_inc("web_search_count", 1)

    results = await _ddg_search(query, n=6)
    if not results:
        return SearchResponse(
            answer="No useful web results were found.",
            citations=[],
            has_results=False,
            query=query,
        )

    citations: list[Citation] = []
    for idx, r in enumerate(results[:6]):
        url = r.get("href") or r.get("url") or ""
        title = r.get("title") or url
        body = r.get("body") or r.get("snippet") or ""
        if not url:
            continue
        try:
            src_label = urlparse(url).netloc.replace("www.", "")
        except Exception:
            src_label = "Web"
        citations.append(
            Citation(
                title=title,
                source_label=src_label,
                url=url,
                pdf_url=url if url.lower().endswith(".pdf") else None,
                section_ref=None,
                excerpt=re.sub(r"\s+", " ", body)[:260].strip(),
                score=float(1.0 - idx * 0.1),
            )
        )

    # Short 1-2 sentence synthesis via gpt-4o-mini
    answer = ""
    if openai_client and citations:
        ctx = "\n\n".join(
            f"[{i+1}] {c.title} ({c.source_label})\n{c.excerpt}"
            for i, c in enumerate(citations[:6])
        )
        system = (
            "You summarise web search results in 1-2 short sentences. "
            "Use ONLY the snippets provided. Do not invent. No legal advice. "
            "No headings. Plain prose."
        )
        user = (
            f"Question: {query}\n\nWeb snippets:\n{ctx}\n\n"
            "Brief 1-2 sentence summary:"
        )
        try:
            resp = await openai_client.chat.completions.create(
                model=SUMMARY_MODEL,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                temperature=0.3,
                max_tokens=120,
            )
            answer = (resp.choices[0].message.content or "").strip()
            await _metric_inc("openai_chat_calls", 1)
            if resp.usage:
                await _metric_inc("openai_chat_tokens", resp.usage.total_tokens)
        except Exception as exc:
            logger.warning("web summary failed: %s", exc)
    if not answer:
        answer = "Top relevant web results are listed below."

    return SearchResponse(
        answer=answer, citations=citations, has_results=True, query=query
    )


# ---------------------------------------------------------------------------
# Admin — hidden system-health dashboard
# ---------------------------------------------------------------------------
@api.post("/admin/auth")
async def admin_auth(req: AuthRequest):
    if req.password.strip() != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid password")
    return {"ok": True, "token": "admin-session"}


@api.get("/admin/health")
async def admin_health():
    metrics = await db.metrics.find_one({"_id": "global"}) or {}
    sources = (
        await db.sources.find({}, {"_id": 0}).sort("created_at", 1).to_list(100)
    )
    total_docs = await db.documents.count_documents({})
    contact_count = await db.contact_submissions.count_documents({})
    chatbot_count = await db.chatbot_submissions.count_documents({})

    # Stale-crawl auto-recovery: if a source has been "crawling" longer than
    # the configured timeout, mark it timeout so the dashboard isn't stuck.
    now_dt = datetime.now(timezone.utc)
    stale_cutoff = now_dt - timedelta(seconds=CRAWL_TIMEOUT_SECONDS + 60)
    for s in sources:
        if s.get("status") != "crawling":
            continue
        started_iso = s.get("current_run_started_at")
        try:
            started_dt = (
                datetime.fromisoformat(started_iso) if started_iso else None
            )
        except Exception:
            started_dt = None
        if started_dt and started_dt < stale_cutoff:
            run_id = s.get("current_run_id")
            if run_id:
                await db.documents.delete_many(
                    {"source_id": s["id"], "crawl_run_id": run_id}
                )
            await db.sources.update_one(
                {"id": s["id"]},
                {
                    "$set": {
                        "status": "timeout",
                        "last_error": (
                            f"timeout after {CRAWL_TIMEOUT_SECONDS}s — "
                            "previous index preserved"
                        ),
                        "last_finished_at": now_dt.isoformat(),
                    },
                    "$unset": {
                        "current_run_id": "",
                        "current_run_started_at": "",
                    },
                },
            )
            s["status"] = "timeout"
            s["last_error"] = "timeout — previous index preserved"

    last_crawl_at = None
    any_error = False
    is_crawling = False
    for s in sources:
        st = s.get("status")
        if st == "crawling":
            is_crawling = True
        if st in ("error", "timeout"):
            any_error = True
        lc = s.get("last_crawled_at")
        if lc and (last_crawl_at is None or lc > last_crawl_at):
            last_crawl_at = lc

    crawl_status = (
        "in_progress" if is_crawling else ("partial" if any_error else "ok")
    )

    # Live per-source doc count from `documents` (excluding any in-flight new
    # docs from a currently-running crawl, so admin sees the previous good
    # number rather than a partial intermediate).
    sources_detail = []
    for s in sources:
        sid = s["id"]
        run_id = s.get("current_run_id")
        live_filter = {"source_id": sid}
        if run_id:
            live_filter["crawl_run_id"] = {"$ne": run_id}
        live_count = await db.documents.count_documents(live_filter)
        # If old-style docs (no crawl_run_id field at all) live_count works.
        # If a fresh crawl has wiped them out historically, this is 0.
        sources_detail.append(
            {
                "id": sid,
                "label": s.get("label", ""),
                "url": s["url"],
                "status": s.get("status", "idle"),
                "pages_indexed": (
                    live_count if live_count else s.get("pages_indexed", 0)
                ),
                "indexed": live_count,
                "sections_indexed": s.get("sections_indexed", 0),
                "pdfs_indexed": s.get("pdfs_indexed", 0),
                "last_crawled_at": s.get("last_crawled_at"),
                "last_error": s.get("last_error", ""),
                "current_run_started_at": s.get("current_run_started_at"),
            }
        )

    return {
        "uptime_started_at": SERVER_STARTED_AT,
        "indexed_documents": total_docs,
        "sources_count": len(sources),
        "sources_detail": sources_detail,
        "crawl_status": crawl_status,
        "last_crawl_at": last_crawl_at,
        "next_scheduled_crawl": _next_scheduled_run_iso(),
        "crawl_interval_days": CRAWL_INTERVAL_DAYS,
        "crawl_timeout_seconds": CRAWL_TIMEOUT_SECONDS,
        "metrics": {
            "openai_embed_calls": int(metrics.get("openai_embed_calls", 0)),
            "openai_embed_tokens": int(metrics.get("openai_embed_tokens", 0)),
            "openai_chat_calls": int(metrics.get("openai_chat_calls", 0)),
            "openai_chat_tokens": int(metrics.get("openai_chat_tokens", 0)),
            "vineyard_search_count": int(
                metrics.get("vineyard_search_count", 0)
            ),
            "web_search_count": int(metrics.get("web_search_count", 0)),
            "chatbot_submissions": chatbot_count,
            "contact_submissions": contact_count,
        },
    }


@api.post("/admin/sources/{source_id}/reset")
async def admin_reset_source(source_id: str):
    """Force-clear a stuck source. Drops any in-flight (run-tagged) chunks but
    preserves the last successfully indexed dataset."""
    src = await db.sources.find_one({"id": source_id}, {"_id": 0})
    if not src:
        raise HTTPException(status_code=404, detail="Source not found")
    run_id = src.get("current_run_id")
    if run_id:
        await db.documents.delete_many(
            {"source_id": source_id, "crawl_run_id": run_id}
        )
    live_count = await db.documents.count_documents({"source_id": source_id})
    new_status = "done" if live_count > 0 else "idle"
    await db.sources.update_one(
        {"id": source_id},
        {
            "$set": {
                "status": new_status,
                "last_error": "manual reset by admin",
                "last_finished_at": datetime.now(timezone.utc).isoformat(),
            },
            "$unset": {"current_run_id": "", "current_run_started_at": ""},
        },
    )
    return {"ok": True, "status": new_status, "indexed": live_count}


@api.post("/admin/seed/rebuild")
async def admin_rebuild_seed():
    """Atlas-backed deployment — seed bundles are no longer used.

    The index lives permanently on MongoDB Atlas. This endpoint is kept
    only so older clients calling it get a clear no-op response.
    """
    return {
        "ok": True,
        "deprecated": True,
        "message": (
            "Seed bundles are no longer used. The index is stored on "
            "MongoDB Atlas and persists across all deployments."
        ),
    }


# ---------------------------------------------------------------------------
# Wire up
# ---------------------------------------------------------------------------
app.include_router(api)

# ---------------------------------------------------------------------------
# WoodChat — premium messaging module mounted under /api/woodchat
# ---------------------------------------------------------------------------
from woodchat import build_woodchat_router  # noqa: E402

woodchat_router = build_woodchat_router(
    db, openai_client=openai_client, summary_model=SUMMARY_MODEL
)
app.include_router(woodchat_router, prefix="/api")

# ---------------------------------------------------------------------------
# EON — standalone premium AI assistant mounted under /api/eon-app
# ---------------------------------------------------------------------------
from eon_app import build_eon_router  # noqa: E402

eon_router = build_eon_router(
    db, openai_client=openai_client, chat_model=SUMMARY_MODEL
)
app.include_router(eon_router, prefix="/api")

# ---------------------------------------------------------------------------
# Research Mode — advanced intelligence workspace mounted under /api/research
# ---------------------------------------------------------------------------
from research import build_research_router  # noqa: E402

research_router = build_research_router(
    db, openai_client=openai_client, summary_model=SUMMARY_MODEL
)
app.include_router(research_router, prefix="/api")

# ---------------------------------------------------------------------------
# Google OAuth — white-label sign-in for EON and WoodX
# ---------------------------------------------------------------------------
from google_auth import build_google_auth_router  # noqa: E402

google_auth_router = build_google_auth_router(db)
app.include_router(google_auth_router, prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


DEFAULT_VINEYARD_SOURCES = [
    ("https://vineyard.municipalcodeonline.com/", "Municipal Code"),
    ("https://www.vineyardutah.gov/", "Vineyard Utah (Official)"),
    ("https://vineyardut.portal.civicclerk.com/", "CivicClerk (Meetings)"),
    (
        "https://cms3.revize.com/revize/vineyard/Departmnts/"
        "Vineyard%20RDA%20past%20meetings%20index.xlsx",
        "RDA Past Meetings Index",
    ),
]

# Sources that should NEVER be auto-crawled or auto-seeded. If any are already
# present in the DB (from an older seed bundle), startup will delete them so
# opening /vineyard never stalls on them.
BANNED_SOURCE_HOSTS = (
    "example.org",
    "example.com",
)


# ---------------------------------------------------------------------------
# Pre-built index hydration  (runs BEFORE seed_default_sources so the bundled
# source UUIDs match the bundled documents' source_id refs)
# ---------------------------------------------------------------------------
SEED_DIR = ROOT_DIR / "seed_data"
SEED_COLLECTIONS = ("documents", "sources", "metrics", "index_meta")


def _decode_bson_gz(path: Path):
    """Decode all BSON documents from a gzipped mongodump file."""
    import gzip
    import bson

    with gzip.open(path, "rb") as fh:
        data = fh.read()
    return bson.decode_all(data)


@app.on_event("startup")
async def hydrate_index_from_seed():
    """Atlas-backed deployment — the index lives permanently on the
    external MongoDB Atlas cluster. Seed-bundle hydration is intentionally
    a no-op so we never overwrite the cloud archive on container start.
    """
    return


@app.on_event("startup")
async def seed_default_sources():
    for url, label in DEFAULT_VINEYARD_SOURCES:
        existing = await db.sources.find_one({"url": url}, {"_id": 0})
        if existing:
            # keep label in sync with config if the admin hasn't customised it
            if not existing.get("label"):
                await db.sources.update_one(
                    {"id": existing["id"]}, {"$set": {"label": label}}
                )
            continue
        rec = SourceRecord(url=url, label=label)
        await db.sources.insert_one(rec.model_dump())
        logger.info("Seeded default source: %s (%s)", url, label)


@app.on_event("startup")
async def reset_stuck_crawls_on_startup():
    """Any source left in `crawling` from a previous boot is by definition
    orphaned (the previous process is gone). Reset its status so it never
    blocks the UI on a fresh production deploy. Also purge any source on the
    banned list so bad seeds from older snapshots do not reappear."""
    try:
        # 1) Purge banned sources entirely (e.g. Utah Transparency)
        banned = await db.sources.find(
            {"url": {"$regex": "|".join(BANNED_SOURCE_HOSTS)}},
            {"_id": 0, "id": 1, "url": 1},
        ).to_list(50)
        for b in banned:
            await db.documents.delete_many({"source_id": b["id"]})
            await db.sources.delete_one({"id": b["id"]})
        if banned:
            logger.info(
                "purged %d banned source(s) on startup: %s",
                len(banned),
                ", ".join(b["url"] for b in banned),
            )

        # 2) Reset stuck crawling rows
        stuck = await db.sources.find(
            {"status": "crawling"}, {"_id": 0}
        ).to_list(50)
        for s in stuck:
            run_id = s.get("current_run_id")
            if run_id:
                await db.documents.delete_many(
                    {"source_id": s["id"], "crawl_run_id": run_id}
                )
            live_count = await db.documents.count_documents(
                {"source_id": s["id"]}
            )
            new_status = "done" if live_count > 0 else "idle"
            await db.sources.update_one(
                {"id": s["id"]},
                {
                    "$set": {
                        "status": new_status,
                        "last_error": "auto-reset on server start",
                    },
                    "$unset": {
                        "current_run_id": "",
                        "current_run_started_at": "",
                    },
                },
            )
        if stuck:
            logger.info(
                "auto-reset %d stuck source(s) from previous boot", len(stuck)
            )
    except Exception as exc:
        logger.warning("stuck-crawl reset skipped: %s", exc)


@app.on_event("startup")
async def start_scheduler():
    """Auto-recrawl is INTENTIONALLY DISABLED.

    The Vineyard archive is a stored, locked dataset. Rebuilds happen ONLY
    when an admin triggers `/api/vineyard/admin/rebuild-index` and then
    promotes the new build via `/api/vineyard/admin/lock-index`. Nothing on
    the user side — page load, deploy, scheduler — touches the archive.
    """
    logger.info("vineyard auto-scheduler disabled — admin-only rebuild")
    return


@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        if scheduler.running:
            scheduler.shutdown(wait=False)
    except Exception:
        pass
    client.close()
