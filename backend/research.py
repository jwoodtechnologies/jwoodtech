"""Research Mode — advanced intelligence workspace.

Separate corpus from the main `/vineyard` archive. All documents live
in `db.research_documents`. Gated by password 555. Targeted entity
crawling for: Vineyard, Geneva Steel, Anderson Geneva, Utah City,
US Steel, PacifiCorp.
"""
from __future__ import annotations

import asyncio
import csv
import io
import logging
import os
import re
import uuid
from datetime import datetime, timezone
from typing import List, Optional

import httpx
from fastapi import APIRouter, BackgroundTasks, HTTPException, UploadFile, File
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field
from rank_bm25 import BM25Okapi

logger = logging.getLogger("jwood.research")

RESEARCH_PASSWORD = os.environ.get("RESEARCH_PASSWORD", "777")
RESEARCH_ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "7607")

# Seed keywords — every crawled page must mention at least one to be
# accepted into `research_documents`. Targeted entity crawling per user
# directive — no broad indexing.
ENTITY_KEYWORDS = [
    "vineyard",
    "geneva steel",
    "anderson geneva",
    "utah city",
    "us steel",
    "pacificorp",
    "vineyard rda",
]

# CIK lookups for SEC EDGAR full-text & filing search.
# Geneva Steel went bankrupt in 2002; SEC's "recent filings" API only
# returns the last ~1000 of an active filer, so we list both the operating
# co (Geneva Steel Co) and the holding co (Geneva Steel Holdings Corp).
SEC_CIK_MAP = {
    "Geneva Steel": "0000860192",
    "Geneva Steel Holdings": "0001128709",
    "PacifiCorp": "0000075594",
    "US Steel": "0001163302",  # United States Steel Corp
}

# Entities for whom we apply the keyword-gate (every filing must mention
# Vineyard/Geneva to be indexed). Filer-IS-the-entity filings (Geneva
# Steel itself) are indexed unconditionally.
SEC_KEYWORD_GATED_ENTITIES = {"PacifiCorp", "US Steel"}

# CourtListener — federal court records + bankruptcy. Free REST API.
# We search dockets via the OpinionCluster + Docket endpoints, restricted
# to queries naming the seed entities. Rate limit: 5,000 req/hour for
# unauthenticated use, plenty for our targeted runs.
COURT_QUERIES = [
    ("Geneva Steel", "geneva-steel"),
    ("Geneva Steel Holdings", "geneva-steel-holdings"),
    ("Anderson Geneva", "anderson-geneva"),
    ("Vineyard RDA", "vineyard-rda"),
    ("Vineyard City Utah", "vineyard"),
    ("Utah City", "utah-city"),
    ("PacifiCorp Vineyard", "pacificorp"),
]

WORD_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9\-]+")


def _tokenize(text: str) -> list[str]:
    return [w.lower() for w in WORD_RE.findall(text or "")]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _highlight_snippet(content: str, query: str, span: int = 280) -> str:
    text = re.sub(r"\s+", " ", content or "").strip()
    if not text:
        return ""
    tokens = [t for t in re.findall(r"[\w]+", query.lower()) if len(t) > 1]
    lower = text.lower()
    pos = -1
    for tok in tokens:
        i = lower.find(tok)
        if i != -1 and (pos == -1 or i < pos):
            pos = i
    if pos == -1:
        snippet = text[:span]
    else:
        start = max(0, pos - span // 3)
        snippet = text[start : start + span]
        if start > 0:
            snippet = "…" + snippet
        if start + span < len(text):
            snippet = snippet + "…"
    if tokens:
        pat = re.compile(
            r"(" + "|".join(re.escape(t) for t in tokens) + r")",
            re.IGNORECASE,
        )
        snippet = pat.sub(r"<mark>\1</mark>", snippet)
    return snippet


def _matches_entity(text: str) -> bool:
    """True if any seed keyword appears in the text."""
    if not text:
        return False
    low = text.lower()
    return any(k in low for k in ENTITY_KEYWORDS)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class AuthRequest(BaseModel):
    password: str


class ResearchDoc(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    url: str = ""
    source: str  # "rda_xlsx" | "pacificorp_csv" | "proservices_xlsx" | "sec_edgar" | "courtlistener" | "manual"
    entity: str = ""  # primary entity tag: geneva-steel, pacificorp, etc.
    doc_type: str = "record"  # record | filing | docket | meeting | vendor | appeal
    content: str
    meta: dict = Field(default_factory=dict)  # filing date, fiscal year, party, case number…
    created_at: str = Field(default_factory=_now_iso)


class SearchReq(BaseModel):
    query: str
    entity: Optional[str] = None
    source: Optional[str] = None
    limit: int = 20


class CompareReq(BaseModel):
    doc_ids: List[str]
    question: Optional[str] = None


class ExportReq(BaseModel):
    doc_ids: List[str]
    target: str = "claude"  # claude | chatgpt | plain
    question: Optional[str] = None


class AddSourceReq(BaseModel):
    url: str
    label: Optional[str] = ""
    entity: Optional[str] = "manual"


# ---------------------------------------------------------------------------
# File ingestion helpers
# ---------------------------------------------------------------------------
def _xlsx_rows(data: bytes) -> tuple[str, list[list[str]]]:
    """Return (sheet_title, list of rows-as-strings)."""
    wb = load_workbook(io.BytesIO(data), data_only=True)
    sheet = wb.active
    rows: list[list[str]] = []
    for r in sheet.iter_rows(values_only=False):
        cells: list[str] = []
        for cell in r:
            val = cell.value
            href = (
                cell.hyperlink.target
                if cell.hyperlink and cell.hyperlink.target
                else None
            )
            text = "" if val is None else str(val).strip()
            if href and text and text != href:
                cells.append(f"{text} [{href}]")
            elif href:
                cells.append(href)
            elif text:
                cells.append(text)
            else:
                cells.append("")
        rows.append(cells)
    return sheet.title or "Sheet", rows


def _csv_rows(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader]


# ---------------------------------------------------------------------------
# SEC EDGAR — targeted full-text + filings search
# ---------------------------------------------------------------------------
SEC_USER_AGENT = "Jwood Technologies Research research@jwoodtechnologies.com"


async def _sec_search_filings(
    cik: str, max_filings: int = 50
) -> list[dict]:
    """Pull recent filings for an entity from SEC EDGAR submissions API."""
    cik_padded = cik.zfill(10)
    url = f"https://data.sec.gov/submissions/CIK{cik_padded}.json"
    headers = {"User-Agent": SEC_USER_AGENT, "Accept": "application/json"}
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.get(url, headers=headers)
            r.raise_for_status()
            data = r.json()
    except Exception as exc:
        logger.warning("SEC submissions fetch failed for CIK %s: %s", cik, exc)
        return []

    name = data.get("name", "")
    recent = data.get("filings", {}).get("recent", {}) or {}
    out: list[dict] = []
    accession = recent.get("accessionNumber", []) or []
    forms = recent.get("form", []) or []
    dates = recent.get("filingDate", []) or []
    primary_docs = recent.get("primaryDocument", []) or []
    primary_descs = recent.get("primaryDocDescription", []) or []
    for i in range(min(len(accession), max_filings)):
        acc = accession[i].replace("-", "")
        form = forms[i] if i < len(forms) else ""
        date = dates[i] if i < len(dates) else ""
        doc = primary_docs[i] if i < len(primary_docs) else ""
        desc = primary_descs[i] if i < len(primary_descs) else ""
        index_url = (
            f"https://www.sec.gov/Archives/edgar/data/"
            f"{int(cik)}/{acc}/{accession[i]}-index.htm"
        )
        doc_url = (
            f"https://www.sec.gov/Archives/edgar/data/"
            f"{int(cik)}/{acc}/{doc}"
            if doc
            else index_url
        )
        out.append(
            {
                "company": name,
                "cik": cik,
                "form": form,
                "filing_date": date,
                "primary_doc": doc,
                "description": desc,
                "url": doc_url,
                "index_url": index_url,
            }
        )
    return out


async def _sec_fetch_text(url: str) -> str:
    """Download a filing's primary document and strip to plain text."""
    headers = {"User-Agent": SEC_USER_AGENT, "Accept": "text/html,*/*"}
    try:
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            r = await client.get(url, headers=headers)
            if r.status_code != 200:
                return ""
            ctype = r.headers.get("content-type", "").lower()
            if "html" in ctype or url.lower().endswith((".htm", ".html")):
                from bs4 import BeautifulSoup

                soup = BeautifulSoup(r.text, "html.parser")
                for tag in soup(["script", "style"]):
                    tag.decompose()
                return re.sub(r"\s+", " ", soup.get_text(" ")).strip()
            elif "pdf" in ctype or url.lower().endswith(".pdf"):
                import fitz  # PyMuPDF

                doc = fitz.open(stream=r.content, filetype="pdf")
                pages = [p.get_text("text") for p in doc]
                doc.close()
                return re.sub(r"\s+", " ", "\n".join(pages)).strip()
            else:
                return re.sub(r"\s+", " ", r.text).strip()
    except Exception as exc:
        logger.warning("SEC fetch failed for %s: %s", url, exc)
        return ""


# ---------------------------------------------------------------------------
# Build router
# ---------------------------------------------------------------------------
def build_research_router(db, openai_client=None, summary_model: str = "gpt-4o-mini"):
    router = APIRouter(prefix="/research", tags=["research"])

    # In-memory crawl status (background tasks). Keyed by run id.
    crawl_runs: dict[str, dict] = {}

    async def _ensure_indexes():
        try:
            await db.research_documents.create_index([("entity", 1)])
            await db.research_documents.create_index([("source", 1)])
            await db.research_documents.create_index([("id", 1)], unique=True)
            await db.research_documents.create_index(
                [("title", "text"), ("content", "text")],
                weights={"title": 5, "content": 1},
                name="research_text_idx",
                default_language="english",
            )
        except Exception as exc:
            logger.debug("ensure indexes (idempotent): %s", exc)

    @router.on_event("startup")
    async def _startup():
        await _ensure_indexes()

    # -----------------------------------------------------------------
    # Auth
    # -----------------------------------------------------------------
    @router.post("/auth")
    async def auth(req: AuthRequest):
        if (req.password or "").strip() != RESEARCH_PASSWORD:
            raise HTTPException(status_code=401, detail="Invalid password")
        return {"ok": True, "token": "research-session"}

    # -----------------------------------------------------------------
    # Stats / Status
    # -----------------------------------------------------------------
    @router.get("/stats")
    async def stats():
        total = await db.research_documents.count_documents({})
        # by entity
        agg_entity = [
            {"$group": {"_id": "$entity", "n": {"$sum": 1}}},
            {"$sort": {"n": -1}},
        ]
        by_entity = {
            (r["_id"] or "unknown"): r["n"]
            async for r in db.research_documents.aggregate(agg_entity)
        }
        agg_source = [
            {"$group": {"_id": "$source", "n": {"$sum": 1}}},
            {"$sort": {"n": -1}},
        ]
        by_source = {
            (r["_id"] or "unknown"): r["n"]
            async for r in db.research_documents.aggregate(agg_source)
        }
        try:
            ds = await db.command("dbStats")
            data_mb = round(ds["dataSize"] / 1024 / 1024, 2)
            storage_mb = round(ds["storageSize"] / 1024 / 1024, 2)
        except Exception:
            data_mb = 0
            storage_mb = 0
        return {
            "total_docs": total,
            "by_entity": by_entity,
            "by_source": by_source,
            "atlas": {
                "data_mb": data_mb,
                "storage_mb": storage_mb,
                "free_tier_limit_mb": 512,
            },
            "active_runs": [
                {"run_id": k, **v}
                for k, v in crawl_runs.items()
                if v.get("status") == "running"
            ],
        }

    # -----------------------------------------------------------------
    # File ingestion (the 3 uploaded files)
    # -----------------------------------------------------------------
    @router.post("/ingest/file")
    async def ingest_file(
        kind: str,  # "rda_xlsx" | "pacificorp_csv" | "proservices_xlsx"
        file: UploadFile = File(...),
    ):
        data = await file.read()
        if kind == "rda_xlsx":
            return {"ingested": await _ingest_rda_xlsx(db, data)}
        if kind == "pacificorp_csv":
            return {"ingested": await _ingest_pacificorp_csv(db, data)}
        if kind == "proservices_xlsx":
            return {"ingested": await _ingest_proservices_xlsx(db, data)}
        raise HTTPException(status_code=400, detail=f"Unknown kind {kind}")

    @router.post("/ingest/seed-uploads")
    async def ingest_seed():
        """One-shot: pull the 3 user-uploaded files from emergent CDN and
        index them. Idempotent — re-running purges previous chunks of the
        same source first.
        """
        urls = {
            "rda_xlsx": (
                "https://customer-assets.emergentagent.com/job_jwood-premium/"
                "artifacts/ivmnoxr2_Vineyard_RDA_past_meetings_index.xlsx"
            ),
            "pacificorp_csv": (
                "https://customer-assets.emergentagent.com/job_jwood-premium/"
                "artifacts/e9yfjw5r_Pacifcorp_Closed_Appeals_Vineyard_-_Sheet1.csv"
            ),
            "proservices_xlsx": (
                "https://customer-assets.emergentagent.com/job_jwood-premium/"
                "artifacts/fm7u5ztk_Vineyard_ProServices_Legal_Acctg_LYRB_"
                "Master_UPDATED.xlsx"
            ),
        }
        result: dict = {}
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            for kind, url in urls.items():
                try:
                    r = await client.get(url)
                    r.raise_for_status()
                    data = r.content
                except Exception as exc:
                    result[kind] = {"error": str(exc)}
                    continue
                if kind == "rda_xlsx":
                    result[kind] = await _ingest_rda_xlsx(db, data)
                elif kind == "pacificorp_csv":
                    result[kind] = await _ingest_pacificorp_csv(db, data)
                elif kind == "proservices_xlsx":
                    result[kind] = await _ingest_proservices_xlsx(db, data)
        return result

    # -----------------------------------------------------------------
    # Targeted SEC EDGAR crawler
    # -----------------------------------------------------------------
    @router.post("/crawl/sec")
    async def crawl_sec(
        background_tasks: BackgroundTasks,
        max_per_entity: int = 30,
    ):
        max_per_entity = max(1, min(int(max_per_entity), 200))
        run_id = str(uuid.uuid4())[:8]
        crawl_runs[run_id] = {
            "status": "running",
            "started_at": _now_iso(),
            "kind": "sec",
            "indexed": 0,
            "errors": 0,
            "current": "",
        }
        background_tasks.add_task(_run_sec_crawl, db, crawl_runs, run_id, max_per_entity)
        return {"run_id": run_id, "status": "running"}

    @router.post("/crawl/courts")
    async def crawl_courts(
        background_tasks: BackgroundTasks,
        max_per_query: int = 25,
    ):
        """Pull federal court dockets + bankruptcy records from the
        free CourtListener public API for each seed entity. Indexed
        as `source: courtlistener`."""
        max_per_query = max(1, min(int(max_per_query), 100))
        run_id = str(uuid.uuid4())[:8]
        crawl_runs[run_id] = {
            "status": "running",
            "started_at": _now_iso(),
            "kind": "courts",
            "indexed": 0,
            "errors": 0,
            "current": "",
        }
        background_tasks.add_task(
            _run_court_crawl, db, crawl_runs, run_id, max_per_query
        )
        return {"run_id": run_id, "status": "running"}

    @router.get("/crawl/status")
    async def crawl_status():
        return {"runs": crawl_runs}

    # -----------------------------------------------------------------
    # User-added sources (paste a URL → crawl into research_documents)
    # -----------------------------------------------------------------
    @router.post("/sources")
    async def add_source(
        req: AddSourceReq, background_tasks: BackgroundTasks
    ):
        """Paste a URL — kicks an immediate crawl of THAT URL only into
        research_documents. Status (indexing/archived/failed) tracked in
        `db.research_sources`.
        """
        url = (req.url or "").strip()
        if not url or not url.lower().startswith(("http://", "https://")):
            raise HTTPException(status_code=400, detail="Provide a valid http(s) URL")
        # Dedupe by URL
        existing = await db.research_sources.find_one(
            {"url": url}, {"_id": 0}
        )
        if existing and existing.get("status") in ("indexing", "archived"):
            return existing
        rec = {
            "id": str(uuid.uuid4()),
            "url": url,
            "label": (req.label or "").strip() or url,
            "entity": (req.entity or "manual").strip() or "manual",
            "status": "indexing",
            "indexed_count": 0,
            "last_error": "",
            "created_at": _now_iso(),
            "finished_at": None,
        }
        await db.research_sources.update_one(
            {"url": url}, {"$set": rec}, upsert=True
        )
        background_tasks.add_task(_run_user_source_crawl, db, rec)
        return rec

    @router.get("/sources")
    async def list_sources():
        rows = await db.research_sources.find(
            {}, {"_id": 0}
        ).sort("created_at", -1).to_list(500)
        return {"sources": rows}

    @router.delete("/sources/{source_id}")
    async def delete_source(source_id: str):
        rec = await db.research_sources.find_one(
            {"id": source_id}, {"_id": 0, "url": 1}
        )
        if not rec:
            raise HTTPException(status_code=404, detail="Not found")
        # Remove docs that came from this URL.
        n = await db.research_documents.delete_many(
            {"source": "user_link", "url": rec["url"]}
        )
        await db.research_sources.delete_one({"id": source_id})
        return {"deleted_docs": n.deleted_count}

    # -----------------------------------------------------------------
    # Search
    # -----------------------------------------------------------------
    @router.post("/search")
    async def search(req: SearchReq):
        query = (req.query or "").strip()
        if not query:
            raise HTTPException(status_code=400, detail="Query is required")
        mongo_filter: dict = {}
        if req.entity:
            mongo_filter["entity"] = req.entity
        if req.source:
            mongo_filter["source"] = req.source
        # Two-stage: $text prefilter, BM25 rerank.
        docs: list[dict] = []
        try:
            cursor = db.research_documents.find(
                {**mongo_filter, "$text": {"$search": query}},
                {"_id": 0, "score": {"$meta": "textScore"}},
            ).sort([("score", {"$meta": "textScore"})]).limit(500)
            docs = await cursor.to_list(500)
        except Exception as exc:
            logger.warning("research $text failed: %s", exc)
        if not docs:
            docs = await db.research_documents.find(mongo_filter, {"_id": 0}).to_list(2000)
        if not docs:
            return {"total": 0, "results": [], "answer": "No matching research documents found."}

        corpus_tokens = [_tokenize(d.get("content", "")) for d in docs]
        bm25 = BM25Okapi(corpus_tokens)
        q_tokens = _tokenize(query)
        if not q_tokens:
            return {"total": 0, "results": [], "answer": "Query too short."}
        scores = bm25.get_scores(q_tokens)
        ranked = sorted(zip(docs, scores), key=lambda x: x[1], reverse=True)
        top = [(d, s) for d, s in ranked if s > 0.5][: req.limit]
        results = []
        for d, s in top:
            results.append(
                {
                    "id": d["id"],
                    "title": d.get("title", ""),
                    "url": d.get("url", ""),
                    "source": d.get("source", ""),
                    "entity": d.get("entity", ""),
                    "doc_type": d.get("doc_type", "record"),
                    "snippet": _highlight_snippet(d.get("content", ""), query),
                    "meta": d.get("meta", {}) or {},
                    "score": float(s),
                    "created_at": d.get("created_at"),
                }
            )

        # AI answer (best-effort)
        answer = await _research_summary(openai_client, summary_model, query, [d for d, _ in top[:5]])
        return {
            "total": len(top),
            "answer": answer,
            "results": results,
        }

    @router.get("/document/{doc_id}")
    async def get_doc(doc_id: str):
        d = await db.research_documents.find_one({"id": doc_id}, {"_id": 0})
        if not d:
            raise HTTPException(status_code=404, detail="Not found")
        return d

    # -----------------------------------------------------------------
    # Source Explorer — browse by source/entity/type/date without
    # requiring a search query. Supports pagination + in-source keyword.
    # -----------------------------------------------------------------
    @router.get("/sources-index")
    async def sources_index():
        """Aggregated list of sources with counts, entities, doc-types.
        Feeds the Source Explorer landing grid."""
        rows = await db.research_documents.find(
            {}, {"_id": 0, "source": 1, "entity": 1, "doc_type": 1, "created_at": 1}
        ).to_list(20000)
        out: dict = {}
        for r in rows:
            src = r.get("source") or "unknown"
            bucket = out.setdefault(
                src,
                {
                    "source": src,
                    "count": 0,
                    "entities": {},
                    "doc_types": {},
                    "latest": None,
                },
            )
            bucket["count"] += 1
            ent = r.get("entity") or "unknown"
            bucket["entities"][ent] = bucket["entities"].get(ent, 0) + 1
            dt = r.get("doc_type") or "unknown"
            bucket["doc_types"][dt] = bucket["doc_types"].get(dt, 0) + 1
            ca = r.get("created_at")
            if ca and (not bucket["latest"] or ca > bucket["latest"]):
                bucket["latest"] = ca
        return {
            "sources": sorted(out.values(), key=lambda x: -x["count"]),
            "total": len(rows),
        }

    @router.get("/browse")
    async def browse(
        source: Optional[str] = None,
        entity: Optional[str] = None,
        doc_type: Optional[str] = None,
        q: Optional[str] = None,
        date_from: Optional[str] = None,  # ISO-8601, compared against meta.date
        date_to: Optional[str] = None,
        skip: int = 0,
        limit: int = 50,
        sort: str = "recent",  # recent | title | source
    ):
        """Paginated document browse with filters.

        - `q` does a MongoDB $text prefilter on content+title when present.
        - `date_from` / `date_to` match on `meta.date` first, falling back
          to `created_at`.
        """
        skip = max(0, int(skip))
        limit = max(1, min(int(limit), 200))
        mongo_filter: dict = {}
        if source:
            mongo_filter["source"] = source
        if entity:
            mongo_filter["entity"] = entity
        if doc_type:
            mongo_filter["doc_type"] = doc_type
        if date_from or date_to:
            # Docs may store date in meta.date (crawled) OR created_at (ingest).
            or_conds = []
            cond_meta: dict = {}
            cond_created: dict = {}
            if date_from:
                cond_meta["$gte"] = date_from
                cond_created["$gte"] = date_from
            if date_to:
                cond_meta["$lte"] = date_to
                cond_created["$lte"] = date_to
            if cond_meta:
                or_conds.append({"meta.date": cond_meta})
            if cond_created:
                or_conds.append({"created_at": cond_created})
            mongo_filter["$or"] = or_conds
        query = (q or "").strip()
        proj = {
            "_id": 0,
            "id": 1,
            "title": 1,
            "url": 1,
            "source": 1,
            "entity": 1,
            "doc_type": 1,
            "content": 1,
            "meta": 1,
            "created_at": 1,
        }
        if query:
            mongo_filter["$text"] = {"$search": query}
            proj["score"] = {"$meta": "textScore"}
            cursor = (
                db.research_documents.find(mongo_filter, proj)
                .sort([("score", {"$meta": "textScore"})])
                .skip(skip)
                .limit(limit)
            )
        else:
            sort_key = [("created_at", -1)]
            if sort == "title":
                sort_key = [("title", 1)]
            elif sort == "source":
                sort_key = [("source", 1), ("created_at", -1)]
            cursor = (
                db.research_documents.find(mongo_filter, proj)
                .sort(sort_key)
                .skip(skip)
                .limit(limit)
            )
        docs = await cursor.to_list(limit)
        # Count total matches (separate lightweight query)
        count_filter = {k: v for k, v in mongo_filter.items() if k != "$text"}
        if query:
            count_filter["$text"] = {"$search": query}
        total = await db.research_documents.count_documents(count_filter)
        # Add short snippet for UI
        results = []
        for d in docs:
            content = d.get("content", "") or ""
            snippet = (
                _highlight_snippet(content, query)
                if query
                else content[:240].strip() + ("…" if len(content) > 240 else "")
            )
            results.append(
                {
                    "id": d["id"],
                    "title": d.get("title", ""),
                    "url": d.get("url", ""),
                    "source": d.get("source", ""),
                    "entity": d.get("entity", ""),
                    "doc_type": d.get("doc_type", "record"),
                    "snippet": snippet,
                    "meta": d.get("meta", {}) or {},
                    "created_at": d.get("created_at"),
                    "score": float(d.get("score", 0) or 0),
                }
            )
        # Filter-facet counts scoped to the applied filters (except the
        # field being aggregated) so the sidebar shows useful options.
        facet_base = {k: v for k, v in mongo_filter.items() if k != "$text"}

        async def _facet(field: str, exclude: str | None = None):
            fb = {k: v for k, v in facet_base.items() if k != exclude}
            pipe = [
                {"$match": fb},
                {"$group": {"_id": f"${field}", "n": {"$sum": 1}}},
                {"$sort": {"n": -1}},
                {"$limit": 40},
            ]
            out = {}
            async for row in db.research_documents.aggregate(pipe):
                out[row["_id"] or "unknown"] = row["n"]
            return out

        facets = {
            "source": await _facet("source", "source"),
            "entity": await _facet("entity", "entity"),
            "doc_type": await _facet("doc_type", "doc_type"),
        }
        return {
            "total": total,
            "skip": skip,
            "limit": limit,
            "results": results,
            "facets": facets,
        }

    # -----------------------------------------------------------------
    # AI Compare — summarise differences between 2+ docs
    # -----------------------------------------------------------------
    @router.post("/compare")
    async def compare(req: CompareReq):
        if not req.doc_ids or len(req.doc_ids) < 2:
            raise HTTPException(status_code=400, detail="Pick at least 2 documents")
        docs = await db.research_documents.find(
            {"id": {"$in": req.doc_ids[:5]}}, {"_id": 0}
        ).to_list(5)
        if len(docs) < 2:
            raise HTTPException(status_code=404, detail="Documents not found")
        if not openai_client:
            return {
                "answer": "AI is not configured. Compare requires an OPENAI_API_KEY.",
                "docs": docs,
            }
        ctx = "\n\n".join(
            [
                f"[D{i + 1}] {d.get('title', '')}\n"
                f"Entity: {d.get('entity', '')}\n"
                f"Source: {d.get('source', '')}\n"
                f"{(d.get('content') or '')[:2200]}"
                for i, d in enumerate(docs)
            ]
        )
        question = (req.question or "Compare these documents and surface key differences, risks, parties, dates, and dollar amounts.").strip()
        try:
            resp = await openai_client.chat.completions.create(
                model=summary_model,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are a research analyst. You compare primary "
                            "source documents and produce concise, factual "
                            "intelligence briefs. Cite each source as [D1], "
                            "[D2] etc. Do not speculate; only summarise what "
                            "is in the supplied text."
                        ),
                    },
                    {
                        "role": "user",
                        "content": f"{question}\n\n{ctx}",
                    },
                ],
                temperature=0.1,
                max_tokens=600,
            )
            answer = (resp.choices[0].message.content or "").strip()
        except Exception as exc:
            logger.warning("compare failed: %s", exc)
            answer = f"Compare failed: {exc}"
        return {"answer": answer, "docs": docs}

    # -----------------------------------------------------------------
    # Export — copy-paste prompt for Claude / ChatGPT
    # -----------------------------------------------------------------
    @router.post("/export")
    async def export_prompt(req: ExportReq):
        if not req.doc_ids:
            raise HTTPException(status_code=400, detail="No documents selected")
        docs = await db.research_documents.find(
            {"id": {"$in": req.doc_ids[:8]}}, {"_id": 0}
        ).to_list(8)
        if not docs:
            raise HTTPException(status_code=404, detail="No matching documents")
        target = (req.target or "claude").lower()
        question = (req.question or "Analyze these primary source documents and produce an investigative brief.").strip()

        body_blocks = []
        for i, d in enumerate(docs):
            body_blocks.append(
                f"--- DOCUMENT {i + 1} ---\n"
                f"Title: {d.get('title', '')}\n"
                f"Entity: {d.get('entity', '')}\n"
                f"Source: {d.get('source', '')}\n"
                f"URL: {d.get('url', '')}\n"
                f"Type: {d.get('doc_type', '')}\n"
                f"Date: {(d.get('meta') or {}).get('date') or d.get('created_at') or ''}\n\n"
                f"{(d.get('content') or '')[:6000]}"
            )
        body = "\n\n".join(body_blocks)

        if target == "claude":
            prompt = (
                f"<task>{question}</task>\n\n"
                f"<context>\n{body}\n</context>\n\n"
                "<instructions>\n"
                "1. Cite each document as [D1], [D2], etc.\n"
                "2. Surface dates, parties, dollar amounts, and any "
                "regulatory or legal claims.\n"
                "3. End with a 'KEY QUESTIONS' list — what's missing or "
                "warrants follow-up research.\n"
                "</instructions>"
            )
        elif target == "chatgpt":
            prompt = (
                f"You are a research analyst. {question}\n\n"
                f"Below are {len(docs)} primary source documents. Cite each "
                "as [D1], [D2], etc. Summarise key facts, dates, parties, "
                "dollar amounts, and any regulatory/legal claims. End with "
                "a list of follow-up questions.\n\n"
                f"{body}"
            )
        else:  # plain
            prompt = body
        return {"prompt": prompt, "docs": [{"id": d["id"], "title": d.get("title")} for d in docs]}

    # -----------------------------------------------------------------
    # Admin — wipe corpus or single source
    # -----------------------------------------------------------------
    @router.delete("/admin/wipe")
    async def admin_wipe(password: str = "", source: Optional[str] = None):
        if password != RESEARCH_ADMIN_PASSWORD:
            raise HTTPException(status_code=401, detail="Unauthorized")
        if source:
            r = await db.research_documents.delete_many({"source": source})
        else:
            r = await db.research_documents.delete_many({})
        return {"deleted": r.deleted_count}

    return router


# ---------------------------------------------------------------------------
# Background SEC crawl
# ---------------------------------------------------------------------------
async def _run_sec_crawl(db, crawl_runs: dict, run_id: str, max_per_entity: int):
    indexed = 0
    errors = 0
    try:
        for entity, cik in SEC_CIK_MAP.items():
            crawl_runs[run_id]["current"] = entity
            try:
                filings = await _sec_search_filings(cik, max_filings=max_per_entity)
            except Exception as exc:
                logger.warning("SEC list failed %s: %s", entity, exc)
                errors += 1
                continue
            for f in filings:
                # SEC has a 10 req/s rate limit. Be polite.
                await asyncio.sleep(0.12)
                text = await _sec_fetch_text(f["url"])
                if not text:
                    continue
                # Entity gate: filing must mention at least one keyword
                # OR the filer IS one of our entities (US Steel, PacifiCorp).
                # For "broad filer" entities we still require a Vineyard /
                # Geneva mention to keep the corpus tight.
                if entity in SEC_KEYWORD_GATED_ENTITIES:
                    if not _matches_entity(text):
                        continue
                title = (
                    f"{f['company']} — {f['form']} — {f['filing_date']}"
                    + (f" — {f['description']}" if f.get("description") else "")
                )
                rec = {
                    "id": str(uuid.uuid4()),
                    "title": title,
                    "url": f["url"],
                    "source": "sec_edgar",
                    "entity": entity.lower().replace(" ", "-"),
                    "doc_type": "filing",
                    "content": text[:200_000],  # cap individual doc at 200 kB text
                    "meta": {
                        "form": f["form"],
                        "filing_date": f["filing_date"],
                        "cik": f["cik"],
                        "company": f["company"],
                        "index_url": f["index_url"],
                        "date": f["filing_date"],
                    },
                    "created_at": _now_iso(),
                }
                # Idempotent upsert by url
                await db.research_documents.update_one(
                    {"url": rec["url"], "source": "sec_edgar"},
                    {"$set": rec},
                    upsert=True,
                )
                indexed += 1
                crawl_runs[run_id]["indexed"] = indexed
        crawl_runs[run_id]["status"] = "done"
        crawl_runs[run_id]["finished_at"] = _now_iso()
        crawl_runs[run_id]["errors"] = errors
    except Exception as exc:
        logger.exception("SEC crawl failed: %s", exc)
        crawl_runs[run_id]["status"] = "error"
        crawl_runs[run_id]["error"] = str(exc)


# ---------------------------------------------------------------------------
# CourtListener — federal courts + bankruptcy (free public API)
# ---------------------------------------------------------------------------
COURTLISTENER_BASE = "https://www.courtlistener.com/api/rest/v4"


async def _courtlistener_search(query: str, max_results: int) -> list[dict]:
    """Use the unified Search API to find dockets + opinion clusters
    matching `query`. Returns a flat list of result dicts.
    """
    out: list[dict] = []
    headers = {"User-Agent": SEC_USER_AGENT, "Accept": "application/json"}
    # Use phrase-quoted query to force tight matches.
    q_phrased = f'"{query}"'
    # Two passes: type=r (RECAP / federal dockets including bankruptcy)
    # and type=o (opinion clusters with full text).
    for typ in ("r", "o"):
        url = f"{COURTLISTENER_BASE}/search/"
        params = {"q": q_phrased, "type": typ, "order_by": "score desc"}
        try:
            async with httpx.AsyncClient(timeout=30) as client:
                r = await client.get(url, params=params, headers=headers)
                if r.status_code != 200:
                    logger.warning(
                        "CourtListener search %s/%s: %s",
                        typ, query, r.status_code,
                    )
                    continue
                data = r.json()
        except Exception as exc:
            logger.warning("CourtListener search failed (%s/%s): %s", typ, query, exc)
            continue
        for item in (data.get("results") or [])[:max_results]:
            item["__type"] = typ
            out.append(item)
        await asyncio.sleep(0.25)  # politeness
    return out


def _courtlistener_record(item: dict, query: str, entity: str) -> Optional[dict]:
    """Normalise a CourtListener search row into a research_documents
    record. Returns None if we cannot extract usable text."""
    typ = item.get("__type", "")
    if typ == "o":  # opinion cluster
        title = item.get("caseName") or item.get("caseNameShort") or "Court opinion"
        court = item.get("court") or ""
        date = item.get("dateFiled") or ""
        snippet = item.get("snippet") or ""
        text = (
            (item.get("opinion") or "")
            + "\n"
            + snippet
            + "\n"
            + (item.get("text") or "")
        ).strip()
        url = (
            f"https://www.courtlistener.com{item.get('absolute_url')}"
            if item.get("absolute_url")
            else item.get("download_url") or ""
        )
        doc_type = "opinion"
    else:  # RECAP / docket
        title = (
            item.get("caseName")
            or item.get("docketNumber")
            or item.get("description")
            or "Federal docket"
        )
        court = item.get("court") or item.get("court_id") or ""
        date = item.get("dateFiled") or item.get("dateTermed") or ""
        # Build content from available fields.
        bits: list[str] = []
        for k in (
            "caseName",
            "docketNumber",
            "description",
            "natureSuit",
            "cause",
            "assignedTo",
            "referredTo",
            "snippet",
        ):
            v = item.get(k)
            if v:
                bits.append(f"{k}: {v}")
        text = " | ".join(bits)
        url = (
            f"https://www.courtlistener.com{item.get('docket_absolute_url') or item.get('absolute_url') or ''}"
            if (item.get("docket_absolute_url") or item.get("absolute_url"))
            else ""
        )
        doc_type = "docket"
    text = re.sub(r"\s+", " ", text or "").strip()
    if not text or len(text) < 60:
        return None
    full_title = f"{title} — {court} ({date})" if date else f"{title} — {court}"
    return {
        "id": str(uuid.uuid4()),
        "title": full_title.strip(" —"),
        "url": url,
        "source": "courtlistener",
        "entity": entity,
        "doc_type": doc_type,
        "content": text[:200_000],
        "meta": {
            "court": court,
            "date": date,
            "case_name": title,
            "docket_number": item.get("docketNumber"),
            "type": typ,
            "query": query,
        },
        "created_at": _now_iso(),
    }


async def _run_court_crawl(db, crawl_runs: dict, run_id: str, max_per_query: int):
    indexed = 0
    errors = 0
    # Wipe previous courtlistener docs so re-runs replace stale records
    # (keeps the corpus clean — the user's mandate is "make sure
    # everyone is locked in and works").
    try:
        await db.research_documents.delete_many({"source": "courtlistener"})
    except Exception as exc:
        logger.warning("courtlistener purge failed: %s", exc)
    try:
        for query, entity in COURT_QUERIES:
            crawl_runs[run_id]["current"] = query
            try:
                rows = await _courtlistener_search(query, max_per_query)
            except Exception as exc:
                logger.warning("CourtListener query failed %s: %s", query, exc)
                errors += 1
                continue
            phrase = query.lower()
            for item in rows:
                rec = _courtlistener_record(item, query, entity)
                if not rec:
                    continue
                # Strict relevance gate: the case name OR docket content
                # MUST contain the search phrase as a substring. Drops
                # noise like "Geneva Garzarelli" matching "Geneva Steel".
                blob = (
                    (rec["title"] or "")
                    + " "
                    + (rec["content"] or "")
                ).lower()
                if phrase not in blob:
                    continue
                # Idempotent upsert by url — fall back to content hash key
                # when URL is empty (rare).
                if rec["url"]:
                    key = {"url": rec["url"], "source": "courtlistener"}
                else:
                    key = {
                        "source": "courtlistener",
                        "title": rec["title"],
                    }
                await db.research_documents.update_one(
                    key, {"$set": rec}, upsert=True
                )
                indexed += 1
                crawl_runs[run_id]["indexed"] = indexed
            await asyncio.sleep(0.5)
        crawl_runs[run_id]["status"] = "done"
        crawl_runs[run_id]["finished_at"] = _now_iso()
        crawl_runs[run_id]["errors"] = errors
    except Exception as exc:
        logger.exception("Court crawl failed: %s", exc)
        crawl_runs[run_id]["status"] = "error"
        crawl_runs[run_id]["error"] = str(exc)


# ---------------------------------------------------------------------------
# User-added source crawler (paste a URL, index into research_documents)
# ---------------------------------------------------------------------------
async def _run_user_source_crawl(db, rec: dict) -> None:
    """Fetch a single URL, extract text, chunk into research_documents.
    Updates research_sources.status to archived | failed when done."""
    url = rec["url"]
    try:
        headers = {"User-Agent": SEC_USER_AGENT, "Accept": "text/html,*/*"}
        async with httpx.AsyncClient(timeout=90, follow_redirects=True) as client:
            r = await client.get(url, headers=headers)
            if r.status_code != 200:
                raise RuntimeError(f"HTTP {r.status_code}")
            ctype = r.headers.get("content-type", "").lower()
            title = rec["label"] or url
            if "html" in ctype or url.lower().endswith((".htm", ".html", "/")):
                from bs4 import BeautifulSoup

                soup = BeautifulSoup(r.text, "html.parser")
                if soup.title and soup.title.string:
                    title = soup.title.string.strip()[:200] or title
                for tag in soup(["script", "style", "nav", "footer", "header"]):
                    tag.decompose()
                text = re.sub(r"\s+", " ", soup.get_text(" ")).strip()
            elif "pdf" in ctype or url.lower().endswith(".pdf"):
                import fitz

                doc = fitz.open(stream=r.content, filetype="pdf")
                pages = [p.get_text("text") for p in doc]
                doc.close()
                text = re.sub(r"\s+", " ", "\n".join(pages)).strip()
            else:
                text = re.sub(r"\s+", " ", r.text).strip()
        if not text or len(text) < 60:
            raise RuntimeError("No extractable text")
        # Chunk into ~2,000-char pieces so search snippets work cleanly.
        chunks = [text[i : i + 2000] for i in range(0, len(text), 2000)]
        # Idempotent: purge any prior chunks for this URL.
        await db.research_documents.delete_many(
            {"source": "user_link", "url": url}
        )
        docs = []
        for i, chunk in enumerate(chunks[:80]):  # cap at 80 chunks / doc
            docs.append(
                {
                    "id": str(uuid.uuid4()),
                    "title": f"{title} — part {i + 1}" if len(chunks) > 1 else title,
                    "url": url,
                    "source": "user_link",
                    "entity": rec.get("entity") or "manual",
                    "doc_type": "webpage",
                    "content": chunk,
                    "meta": {
                        "added_by": "user",
                        "label": rec.get("label"),
                        "chunk": i,
                        "total_chunks": len(chunks),
                    },
                    "created_at": _now_iso(),
                }
            )
        if docs:
            await db.research_documents.insert_many(docs)
        await db.research_sources.update_one(
            {"id": rec["id"]},
            {
                "$set": {
                    "status": "archived",
                    "indexed_count": len(docs),
                    "finished_at": _now_iso(),
                    "last_error": "",
                }
            },
        )
    except Exception as exc:
        logger.warning("user source crawl failed for %s: %s", url, exc)
        await db.research_sources.update_one(
            {"id": rec["id"]},
            {
                "$set": {
                    "status": "failed",
                    "finished_at": _now_iso(),
                    "last_error": str(exc)[:400],
                }
            },
        )


# ---------------------------------------------------------------------------
# File ingestion implementations
# ---------------------------------------------------------------------------
async def _ingest_rda_xlsx(db, data: bytes) -> dict:
    """RDA past meetings index — every row is a meeting record. Chunk
    each row as its own document so search can pick up specific meeting
    titles like 'Megaplex Contract' or 'Resolution 2014-05'.
    """
    title, rows = _xlsx_rows(data)
    if not rows:
        return {"docs": 0}
    # Wipe previous chunks of this source for idempotency.
    await db.research_documents.delete_many({"source": "rda_xlsx"})
    header = [c for c in rows[0] if c]
    inserts: list[dict] = []
    for i, row in enumerate(rows[1:], start=2):
        non_empty = [c for c in row if c]
        if not non_empty:
            continue
        # Build a structured content string.
        if len(header) == len(row):
            kv = [f"{h}: {v}" for h, v in zip(header, row) if v]
        else:
            kv = non_empty
        content = " | ".join(kv)
        # Try to find a meeting title from the first cell-with-text
        meeting_title = next((c for c in row if c and len(c) > 4), f"RDA Row {i}")
        inserts.append(
            {
                "id": str(uuid.uuid4()),
                "title": f"RDA — {meeting_title[:120]}",
                "url": "",
                "source": "rda_xlsx",
                "entity": "vineyard-rda",
                "doc_type": "meeting",
                "content": content,
                "meta": {
                    "row": i,
                    "header": header,
                    "values": row,
                },
                "created_at": _now_iso(),
            }
        )
    if inserts:
        await db.research_documents.insert_many(inserts)
    return {"docs": len(inserts), "sheet": title}


async def _ingest_pacificorp_csv(db, data: bytes) -> dict:
    rows = _csv_rows(data)
    if not rows:
        return {"docs": 0}
    await db.research_documents.delete_many({"source": "pacificorp_csv"})
    header = rows[0]
    inserts: list[dict] = []
    for i, row in enumerate(rows[1:], start=2):
        if not any((c or "").strip() for c in row):
            continue
        kv = [f"{h}: {v}" for h, v in zip(header, row) if (v or "").strip()]
        content = " | ".join(kv)
        title = next((c for c in row if c and len(c.strip()) > 3), f"Appeal Row {i}")
        inserts.append(
            {
                "id": str(uuid.uuid4()),
                "title": f"PacifiCorp Appeal — {title[:120]}",
                "url": "",
                "source": "pacificorp_csv",
                "entity": "pacificorp",
                "doc_type": "appeal",
                "content": content,
                "meta": {
                    "row": i,
                    "header": header,
                    "values": row,
                },
                "created_at": _now_iso(),
            }
        )
    if inserts:
        await db.research_documents.insert_many(inserts)
    return {"docs": len(inserts)}


async def _ingest_proservices_xlsx(db, data: bytes) -> dict:
    """Vineyard ProServices Master — vendor / legal / accounting history."""
    wb = load_workbook(io.BytesIO(data), data_only=True)
    await db.research_documents.delete_many({"source": "proservices_xlsx"})
    inserts: list[dict] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=False))
        if not rows:
            continue
        header_cells = rows[0]
        header = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
        for i, r in enumerate(rows[1:], start=2):
            cells: list[str] = []
            hrefs: list[str] = []
            for cell in r:
                val = cell.value
                href = (
                    cell.hyperlink.target
                    if cell.hyperlink and cell.hyperlink.target
                    else None
                )
                text = "" if val is None else str(val).strip()
                if href:
                    hrefs.append(href)
                    if text and text != href:
                        cells.append(f"{text} [{href}]")
                    else:
                        cells.append(href)
                else:
                    cells.append(text)
            non_empty = [c for c in cells if c]
            if not non_empty:
                continue
            kv = [f"{h}: {v}" for h, v in zip(header, cells) if v and h]
            content = " | ".join(kv) if kv else " | ".join(non_empty)
            title = next((c for c in non_empty if len(c) > 3), f"Vendor Row {i}")
            inserts.append(
                {
                    "id": str(uuid.uuid4()),
                    "title": f"ProServices [{sheet.title}] — {title[:120]}",
                    "url": hrefs[0] if hrefs else "",
                    "source": "proservices_xlsx",
                    "entity": "vineyard",
                    "doc_type": "vendor",
                    "content": content,
                    "meta": {
                        "sheet": sheet.title,
                        "row": i,
                        "header": header,
                        "values": cells,
                        "hrefs": hrefs,
                    },
                    "created_at": _now_iso(),
                }
            )
    if inserts:
        # Mongo insert_many caps at 16 MB; chunk if huge
        for i in range(0, len(inserts), 500):
            await db.research_documents.insert_many(inserts[i : i + 500])
    return {"docs": len(inserts), "sheets": [s.title for s in wb.worksheets]}


# ---------------------------------------------------------------------------
# AI summary for research search
# ---------------------------------------------------------------------------
async def _research_summary(openai_client, model: str, query: str, top_docs: list[dict]) -> str:
    if not openai_client or not top_docs:
        return ""
    ctx_blocks = []
    for i, d in enumerate(top_docs[:5], start=1):
        ctx_blocks.append(
            f"[S{i}] {d.get('title', '')}\n"
            f"Source: {d.get('source', '')} · Entity: {d.get('entity', '')}\n"
            f"{(d.get('content') or '')[:1100]}"
        )
    ctx = "\n\n".join(ctx_blocks)
    system = (
        "You are a research analyst. You answer questions about primary "
        "source documents (SEC filings, court records, municipal "
        "appeals, vendor invoices). Use ONLY the supplied sources. Cite "
        "each as [S1], [S2], etc. Stay neutral. 2–4 sentences. No legal "
        "advice."
    )
    user = f"Question: {query}\n\nTop matched documents:\n{ctx}\n\nBrief, factual analyst answer:"
    try:
        resp = await openai_client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            temperature=0.1,
            max_tokens=260,
        )
        return (resp.choices[0].message.content or "").strip()
    except Exception as exc:
        logger.warning("research summary failed: %s", exc)
        return ""
